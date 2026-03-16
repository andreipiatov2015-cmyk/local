from flask import Flask, render_template, request, jsonify, redirect, url_for, session, g, make_response, send_file
import json
import os
import io
import smtplib
from email.mime.text import MIMEText
import sqlite3
import hashlib
import hmac
import secrets
import datetime
import subprocess
import signal
import time
import threading
import shutil
import re
import uuid
import traceback
import logging
import zipfile
from functools import wraps
from pathlib import Path

import requests


try:
    import openpyxl
except Exception:
    openpyxl = None

# Папка, где лежит server.py
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ------------ Загрузка .env ------------
def load_env(path=".env"):
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    key, value = line.strip().split("=", 1)
                    os.environ[key] = value

# если у тебя .env лежит в /var/www/.env — оставляем
load_env("/var/www/.env")

# ------------ Flask ------------
app = Flask(__name__, template_folder="templates", static_folder="static")

LOG_FILE = os.environ.get("LIVE_SERVER_LOG_FILE", "/var/log/live-server.log")
if not any(isinstance(h, logging.FileHandler) and getattr(h, "baseFilename", "") == LOG_FILE for h in app.logger.handlers):
    try:
        file_handler = logging.FileHandler(LOG_FILE)
        file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)
    except Exception:
        # Если файловый лог не подключился (например, нет прав), используем стандартный логгер Flask.
        app.logger.exception("[logging] failed to initialize file handler for %s", LOG_FILE)


def resolve_app_secret_key():
    env_secret = (
        os.environ.get("APP_SECRET_KEY")
        or os.environ.get("FLASK_SECRET_KEY")
        or os.environ.get("SECRET_KEY")
        or ""
    ).strip()
    if env_secret:
        return env_secret

    secret_key_file = Path(os.environ.get("SECRET_KEY_FILE", os.path.join(BASE_DIR, ".flask_secret_key")))
    if secret_key_file.exists():
        key = secret_key_file.read_text(encoding="utf-8").strip()
        if key:
            return key

    secret_key_file.parent.mkdir(parents=True, exist_ok=True)
    generated = secrets.token_hex(32)
    secret_key_file.write_text(generated, encoding="utf-8")
    try:
        os.chmod(secret_key_file, 0o600)
    except OSError:
        pass
    app.logger.warning("[auth] SECRET_KEY not set in env; generated persistent key at %s", secret_key_file)
    return generated


app.secret_key = resolve_app_secret_key()

# ------------ Константы (абсолютные пути) ------------
DB_FILE = os.path.join(BASE_DIR, "app.db")
ENTRIES_FILE = os.path.join(BASE_DIR, "entries.json")
PRESETS_FILE = os.path.join(BASE_DIR, "presets.json")
VK_SETTINGS_FILE = os.path.join(BASE_DIR, "vk_settings.json")
STREAM_TARGETS_FILE = os.path.join(BASE_DIR, "stream_targets.json")
HLS_STREAM_URL = (os.environ.get("HLS_STREAM_URL") or "").strip()

STORAGE_ROOT = os.environ.get("STORAGE_ROOT", "/var/mount_point/nfv/contest_storage")
APP_DATA_ROOT = Path(os.environ.get("APP_DATA_ROOT", os.path.join(BASE_DIR, "app_data")))
RETENTION_DAYS = 60
MAX_TABLES_PER_USER = 10
DOWNLOAD_RETRIES = 3
MAX_FILENAME_LEN = 150
EXCEL_PREVIEW_LIMIT = 200
YANDEX_VNC_URL = os.environ.get("YANDEX_VNC_URL", "/tables/yandex/vnc/vnc.html?autoconnect=1&resize=scale&show_dot=0")
YANDEX_PROFILE_DIR = os.environ.get("YANDEX_PROFILE_DIR", "/var/mount_point/nfv/contest_storage/yandex_profile")
YANDEX_FORMS_TEST_URL = (os.environ.get("YANDEX_FORMS_TEST_URL") or "").strip()
YANDEX_REFRESH_URL = (os.environ.get("YANDEX_REFRESH_URL") or "https://disk.yandex.ru/client/disk").strip()
YANDEX_CHROMIUM_RESTART_CMD = (os.environ.get("YANDEX_CHROMIUM_RESTART_CMD") or "").strip()
YANDEX_CHROMIUM_OPEN_CMD = (os.environ.get("YANDEX_CHROMIUM_OPEN_CMD") or "").strip()
TABLE_HEADER_TAGS_FILE = os.environ.get("TABLE_HEADER_TAGS_FILE", os.path.join(BASE_DIR, "config", "table_header_tags.json"))
TAG_TYPES_FILE = os.environ.get("TAG_TYPES_FILE", os.path.join(BASE_DIR, "config", "tag_types.json"))

MAPPING_FIELDS = {}
HEADER_ALIASES = {}
BASE_PROGRAM_FIELDS = ["number_title", "participant_fio", "studio_name", "territory", "municipality", "nomination", "age_category", "league"]
DOCUMENT_COLUMN_META = {
    "audio_url": {"title": "Фонограмма", "kind": "file"},
    "receipt_url": {"title": "Квитанция / чек", "kind": "document"},
    "consent_url": {"title": "Согласие на обработку ПД", "kind": "document_list"},
    "application_file": {"title": "Заявка на конкурс", "kind": "document"},
    "presentation_url": {"title": "Презентация", "kind": "document"},
    "video_url": {"title": "Видео", "kind": "link"},
}
FILE_MAPPING_FIELDS = {"audio_url", "presentation_url", "consent_url"}



def load_tag_types_config():
    defaults = {
        "municipality": "grouped_choice",
        "nomination": "grouped_choice",
        "age_category": "grouped_choice",
        "league": "grouped_choice",
        "craft_technique": "grouped_choice",
        "video_url": "links",
        "audio_url": "files",
        "consent_url": "files_or_links",
        "presentation_url": "files_or_links",
        "work_photo": "files",
        "work_sketch": "files",
    }
    try:
        with open(TAG_TYPES_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        types = data.get("types") if isinstance(data, dict) else None
        if isinstance(types, dict):
            cleaned = {str(k).strip(): str(v).strip() for k, v in types.items() if str(k).strip() and str(v).strip()}
            if cleaned:
                return cleaned
    except Exception:
        app.logger.warning("[table_tag_types] failed to load %s, fallback defaults", TAG_TYPES_FILE)
    return defaults


TAG_TYPE_OVERRIDES = load_tag_types_config()

PROGRAM_DOWNLOAD_ROUTE_BY_TAG = {
    "audio_url": "audio",
    "receipt_url": "receipt",
    "presentation_url": "presentation",
}

GROUPED_HEADER_RULES = {
    "territory": [],
    "municipality": ["Муниципалитет", "Территория", "Округ", "Название округа"],
    "nomination": ["Номинация"],
    "age_category": ["Возрастная категория"],
    "league": ["Лига"],
    "craft_technique": ["Техника ДПИ, использованная в работе", "Техника ДПИ"],
}



def _safe_trim(value):
    text = str(value or "")
    text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def load_table_header_tags_config():
    empty = {"version": 1, "tags": []}
    try:
        with open(TABLE_HEADER_TAGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        app.logger.error("[table_tags] config file is missing: %s", TABLE_HEADER_TAGS_FILE)
        return empty
    except json.JSONDecodeError as exc:
        app.logger.error("[table_tags] config file is invalid JSON: %s (%s)", TABLE_HEADER_TAGS_FILE, exc)
        return empty
    except Exception:
        app.logger.exception("[table_tags] failed to load config file: %s", TABLE_HEADER_TAGS_FILE)
        return empty

    if not isinstance(data, dict):
        app.logger.error("[table_tags] config root must be object: %s", TABLE_HEADER_TAGS_FILE)
        return empty

    tags = data.get("tags")
    if not isinstance(tags, list):
        app.logger.error("[table_tags] config 'tags' must be list: %s", TABLE_HEADER_TAGS_FILE)
        return empty

    cleaned_tags = []
    for i, item in enumerate(tags):
        if not isinstance(item, dict):
            app.logger.warning("[table_tags] skipping non-object tag entry at index=%s", i)
            continue
        key = _safe_trim(item.get("key"))
        tag = _safe_trim(item.get("tag"))
        aliases = item.get("aliases")
        if not key or not tag or not isinstance(aliases, list):
            app.logger.warning("[table_tags] skipping invalid tag entry at index=%s key=%s tag=%s", i, key, tag)
            continue
        alias_values = [str(a) for a in aliases if isinstance(a, str)]
        cleaned_tags.append({"key": key, "tag": tag, "aliases": alias_values})

    return {"version": data.get("version", 1), "tags": cleaned_tags}


def init_table_header_tags():
    global MAPPING_FIELDS, HEADER_ALIASES
    config = load_table_header_tags_config()
    MAPPING_FIELDS = {}
    HEADER_ALIASES = {}
    for item in config.get("tags", []):
        key = item["key"]
        MAPPING_FIELDS[key] = item["tag"]
        HEADER_ALIASES[key] = list(item.get("aliases", []))
    app.logger.info("[table_tags] loaded tags=%s aliases=%s from %s", len(MAPPING_FIELDS), sum(len(v) for v in HEADER_ALIASES.values()), TABLE_HEADER_TAGS_FILE)

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "noreply@example.com")

# ------------ Работа с JSON ------------
def load_entries():
    if os.path.exists(ENTRIES_FILE):
        try:
            with open(ENTRIES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_entries(entries):
    os.makedirs(os.path.dirname(ENTRIES_FILE), exist_ok=True)
    with open(ENTRIES_FILE, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=4)

def clear_entries():
    entries.clear()
    save_entries(entries)

def load_presets():
    if os.path.exists(PRESETS_FILE):
        try:
            with open(PRESETS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_presets(presets):
    os.makedirs(os.path.dirname(PRESETS_FILE), exist_ok=True)
    with open(PRESETS_FILE, "w", encoding="utf-8") as f:
        json.dump(presets, f, ensure_ascii=False, indent=4)

entries = load_entries()
presets = load_presets()

# ------------ Работа с БД ------------
def query_db(query, args=(), one=False):
    con = sqlite3.connect(DB_FILE)
    con.row_factory = sqlite3.Row
    cur = con.execute(query, args)
    rv = cur.fetchall()
    con.commit()
    con.close()
    return (rv[0] if rv else None) if one else rv

def init_db():
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()

    # Таблица пользователей
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            password_salt TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer',
            is_verified INTEGER NOT NULL DEFAULT 0,
            verify_token TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    # Таблица мероприятий
    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            address TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS email_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            code TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS table_workspaces (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'new',
            total_count INTEGER NOT NULL DEFAULT 0,
            processed_count INTEGER NOT NULL DEFAULT 0,
            download_cursor_row_id INTEGER,
            progress INTEGER NOT NULL DEFAULT 0,
            last_error TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            yandex_session_json TEXT,
            yandex_status TEXT NOT NULL DEFAULT 'disconnected',
            yandex_last_error TEXT,
            yandex_last_checked_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_mapping_presets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            pattern TEXT NOT NULL,
            match_type TEXT NOT NULL,
            priority INTEGER NOT NULL DEFAULT 100,
            updated_at TEXT NOT NULL,
            UNIQUE(user_id, label, pattern, match_type)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS mapping_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            signature TEXT NOT NULL,
            mapping_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(user_id, signature)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS table_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_id INTEGER NOT NULL,
            number_title TEXT,
            fio TEXT,
            team TEXT,
            unique_key TEXT,
            audio_url TEXT,
            receipt_url TEXT,
            consent_url TEXT,
            application_file TEXT,
            generic_file TEXT,
            video_url TEXT,
            presentation_url TEXT,
            audio_local TEXT,
            receipt_local TEXT,
            consent_local TEXT,
            application_local TEXT,
            generic_local TEXT,
            presentation_local TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS table_program_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_id INTEGER NOT NULL,
            kind TEXT NOT NULL,
            entry_id INTEGER,
            sort_index REAL NOT NULL,
            break_minutes INTEGER,
            notes TEXT,
            is_hidden INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS yandex_connect_sessions (
            connect_id TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            table_id INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'waiting',
            error_message TEXT,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    for stmt in [
        "ALTER TABLE table_workspaces ADD COLUMN excel_headers_json TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN excel_preview_rows_json TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN excel_total_rows INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE table_workspaces ADD COLUMN excel_sheet_name TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN mapping_json TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN total_count INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE table_workspaces ADD COLUMN processed_count INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE table_workspaces ADD COLUMN download_cursor_row_id INTEGER",
        "ALTER TABLE table_workspaces ADD COLUMN last_error TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN yandex_status TEXT NOT NULL DEFAULT 'disconnected'",
        "ALTER TABLE table_workspaces ADD COLUMN yandex_last_error TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN yandex_last_checked_at TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN is_finalized INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE table_workspaces ADD COLUMN finalized_at TEXT",
        "ALTER TABLE table_workspaces ADD COLUMN program_updated_at TEXT",
        "ALTER TABLE table_entries ADD COLUMN row_id INTEGER",
        "ALTER TABLE table_entries ADD COLUMN row_data_json TEXT",
        "ALTER TABLE table_program_items ADD COLUMN notes TEXT",
        "ALTER TABLE table_program_items ADD COLUMN is_hidden INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE table_entries ADD COLUMN conflicts_json TEXT",
        "ALTER TABLE table_entries ADD COLUMN resolved_fields_json TEXT",
        "ALTER TABLE table_entries ADD COLUMN generic_local TEXT",
        "ALTER TABLE table_entries ADD COLUMN application_local TEXT",
        "ALTER TABLE table_entries ADD COLUMN video_url TEXT",
        "ALTER TABLE table_entries ADD COLUMN generic_file TEXT",
        "ALTER TABLE table_entries ADD COLUMN application_file TEXT",
    ]:
        try:
            cur.execute(stmt)
        except sqlite3.OperationalError:
            pass

    con.commit()
    con.close()

# ------------ Пароли и токены ------------
PBKDF2_ITER = 100_000

def make_salt(n=16): return secrets.token_hex(n)

def hash_password(password, salt):
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), PBKDF2_ITER)
    return dk.hex()

def verify_password(password, salt, password_hash):
    return hmac.compare_digest(hash_password(password, salt), password_hash)

def generate_token(n_bytes=24): return secrets.token_urlsafe(n_bytes)

# ------------ Email ------------
def send_email(to_email, subject, body):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASSWORD:
        app.logger.error(
            "[email] SMTP is not configured host=%s user_set=%s password_set=%s to=%s",
            SMTP_HOST,
            bool(SMTP_USER),
            bool(SMTP_PASSWORD),
            to_email,
        )
        raise RuntimeError("SMTP_NOT_CONFIGURED")

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = to_email

    try:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    except Exception as exc:
        # Логируем traceback целиком для диагностики проблем SMTP.
        app.logger.error(
            "[email] send failed to=%s host=%s port=%s error=%s\n%s",
            to_email,
            SMTP_HOST,
            SMTP_PORT,
            exc,
            traceback.format_exc(),
        )
        raise RuntimeError("SMTP_SEND_FAILED") from exc

# ------------ Авторизация ------------
def get_current_user():
    if "user_id" in session:
        row = query_db("SELECT id, username, email, role, is_verified FROM users WHERE id=?", (session["user_id"],), one=True)
        return dict(row) if row else None
    return None

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return f(*a, **kw)
    return wrapper

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*a, **kw):
            user = get_current_user()
            if not user or user["role"] not in roles:
                return "Доступ запрещён", 403
            return f(*a, **kw)
        return wrapper
    return decorator

# ------------ Создание админа по умолчанию ------------
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_EMAIL = "admin@local"
DEFAULT_ADMIN_PASSWORD = "Gfnhbjnjd9"

def get_configured_admin_email():
    return (os.environ.get("ADMIN_EMAIL") or DEFAULT_ADMIN_EMAIL).strip() or DEFAULT_ADMIN_EMAIL


def ensure_admin_exists():
    admin_email = get_configured_admin_email()
    existing = query_db(
        "SELECT id FROM users WHERE username=? OR email=?",
        (DEFAULT_ADMIN_USERNAME, admin_email),
        one=True,
    )
    if existing:
        return existing["id"]

    salt = make_salt()
    pwd_hash = hash_password(DEFAULT_ADMIN_PASSWORD, salt)
    now = datetime.datetime.utcnow().isoformat()
    query_db(
        "INSERT INTO users (username, email, password_hash, password_salt, role, is_verified, created_at, updated_at) VALUES (?, ?, ?, ?, 'admin', 1, ?, ?)",
        (DEFAULT_ADMIN_USERNAME, admin_email, pwd_hash, salt, now, now)
    )
    created = query_db("SELECT id FROM users WHERE username=?", (DEFAULT_ADMIN_USERNAME,), one=True)
    print(f"Создан администратор: {DEFAULT_ADMIN_USERNAME} / {DEFAULT_ADMIN_PASSWORD}")
    return created["id"] if created else None


def migrate_tables_to_current_admin():
    admin_email = get_configured_admin_email()
    admin_candidates = query_db(
        """
        SELECT id, username, email
        FROM users
        WHERE username=? OR email=? OR email=?
        ORDER BY id DESC
        """,
        (DEFAULT_ADMIN_USERNAME, DEFAULT_ADMIN_EMAIL, admin_email),
    )
    if not admin_candidates:
        app.logger.warning("[tables_migration] no admin candidate found; migration skipped")
        return

    current_admin = dict(admin_candidates[0])
    admin_candidate_ids = [row["id"] for row in admin_candidates]
    query_db(
        """
        UPDATE table_workspaces
        SET user_id=?, updated_at=?
        WHERE user_id IN (
            SELECT tw.user_id
            FROM table_workspaces tw
            LEFT JOIN users u ON u.id = tw.user_id
            WHERE u.id IS NULL
        )
        """,
        (current_admin["id"], now_iso()),
    )

    old_existing_admin_ids = [uid for uid in admin_candidate_ids if uid != current_admin["id"]]
    if old_existing_admin_ids:
        placeholders = ",".join(["?"] * len(old_existing_admin_ids))
        query_db(
            f"UPDATE table_workspaces SET user_id=?, updated_at=? WHERE user_id IN ({placeholders})",
            (current_admin["id"], now_iso(), *old_existing_admin_ids),
        )

    app.logger.info(
        "[tables_migration] current_admin_id=%s username=%s email=%s old_admin_ids=%s",
        current_admin["id"],
        current_admin.get("username"),
        current_admin.get("email"),
        old_existing_admin_ids,
    )


def resolve_stream_url():
    if HLS_STREAM_URL:
        # Поддерживаем как абсолютный URL, так и относительный путь из env.
        return HLS_STREAM_URL
    return "/hls/stream.m3u8"

# ------------ Маршруты ------------
@app.route("/")
def index():
    return redirect("/login")

@app.route("/admin")
@login_required
def admin():
    return render_template("admin.html", user=get_current_user(), stream_url=resolve_stream_url())

@app.route("/staff")
@roles_required("admin")
def staff():
    users = query_db("SELECT id, username, email, role, is_verified FROM users ORDER BY id")
    return render_template("staff.html", users=users, stream_url=resolve_stream_url())

@app.route("/broadcast")
def broadcast():
    return render_template("broadcast.html", stream_url=resolve_stream_url())

@app.route("/competition")
def competition():
    return render_template("competition.html", stream_url=resolve_stream_url())

# --- Мероприятия ---
@app.route("/editor")
@login_required
def editor_list():
    return render_template("editor.html", stream_url=resolve_stream_url())

@app.route("/events")
def get_events():
    rows = query_db("SELECT * FROM events ORDER BY id DESC")
    return jsonify([dict(r) for r in rows])

@app.route("/create_event", methods=["POST"])
@login_required
def create_event():
    data = request.json
    query_db(
        "INSERT INTO events (title, address, start_date, end_date) VALUES (?, ?, ?, ?)",
        (data["title"], data["address"], data["start_date"], data["end_date"])
    )
    return jsonify({"status": "ok"})

@app.route("/editor/<int:event_id>")
@login_required
def editor_event(event_id):
    return render_template("admin.html", event_id=event_id, stream_url=resolve_stream_url())

# --- Участники / пресеты ---
@app.route("/entries")
def get_entries():
    return jsonify(entries)

@app.route("/entries/clear", methods=["POST"])
def clear_entries_route():
    clear_entries()
    return jsonify({"status": "ok", "count": 0})

@app.route("/add_entry", methods=["POST"])
def add_entry():
    data = request.json or {}
    new_id = max([entry.get("id", 0) for entry in entries], default=0) + 1
    data["id"] = new_id
    entries.append(data)
    save_entries(entries)
    return jsonify({"message": "Запись добавлена"}), 200

@app.route("/update_entry/<int:id>", methods=["PUT"])
def update_entry(id):
    data = request.json or {}
    for entry in entries:
        if entry.get("id") == id:
            entry.update(data)
            save_entries(entries)
            return jsonify({"message": "Запись обновлена"}), 200
    return jsonify({"error": "Запись не найдена"}), 404

@app.route("/delete_entry/<int:id>", methods=["DELETE"])
def delete_entry(id):
    global entries
    entries = [entry for entry in entries if entry.get("id") != id]
    save_entries(entries)
    return jsonify({"message": "Запись удалена"}), 200

@app.route("/save_all_entries", methods=["POST"])
def save_all_entries():
    data = request.json or []
    entries.clear()
    entries.extend(data)
    save_entries(entries)
    return jsonify({"message": "Все записи сохранены"}), 200

@app.route("/reorder_entries", methods=["POST"])
def reorder_entries():
    data = request.json or []
    entries.clear()
    entries.extend(data)
    save_entries(entries)
    return jsonify({"message": "Порядок обновлён"}), 200

@app.route("/presets")
def get_presets():
    return jsonify(presets)

@app.route("/save_preset", methods=["POST"])
def save_preset():
    data = request.json or {}
    presets.append(data)
    save_presets(presets)
    return jsonify({"message": "Пресет сохранён"}), 200

# --- Управление пользователями ---
@app.route("/staff/change_role/<int:user_id>", methods=["POST"])
@roles_required("admin")
def staff_change_role(user_id):
    new_role = request.form.get("role")
    if new_role not in ["admin", "editor", "viewer"]:
        return "Недопустимая роль", 400
    query_db("UPDATE users SET role=?, updated_at=? WHERE id=?", (new_role, datetime.datetime.utcnow().isoformat(), user_id))
    return redirect(url_for("staff"))

@app.route("/staff/change_password/<int:user_id>", methods=["POST"])
@roles_required("admin")
def staff_change_password(user_id):
    new_password = request.form.get("new_password")
    if not new_password:
        return "Пароль не может быть пустым", 400
    salt = make_salt()
    pwd_hash = hash_password(new_password, salt)
    query_db("UPDATE users SET password_hash=?, password_salt=?, updated_at=? WHERE id=?", (pwd_hash, salt, datetime.datetime.utcnow().isoformat(), user_id))
    return redirect(url_for("staff"))

# --- Логин / Регистрация ---
@app.route("/login", methods=["GET", "POST"])
def login():
    error_message = None
    if request.method == "POST":
        username_or_email = request.form.get("username", "")
        password = request.form.get("password", "")
        row = query_db("SELECT * FROM users WHERE username=? OR email=?", (username_or_email, username_or_email), one=True)
        if not row or not verify_password(password, row["password_salt"], row["password_hash"]):
            error_message = "Неверный логин или пароль"
            return render_template("login.html", error_message=error_message), 401
        if row["is_verified"] != 1:
            error_message = "Подтвердите e-mail."
            return render_template("login.html", error_message=error_message), 403
        session["user_id"] = row["id"]
        return redirect(url_for("admin"))
    return render_template("login.html", error_message=error_message)

@app.route("/logout")
def logout():
    session.pop("user_id", None)
    return redirect(url_for("login"))

@app.route("/register", methods=["GET", "POST"])
def register():
    status_message = None
    error_message = None
    status_code = 200
    if request.method == "POST":
        username = request.form.get("username", "")
        email = request.form.get("email", "")
        password = request.form.get("password", "")
        confirm = request.form.get("confirm", "")
        wants_json = request.is_json or "application/json" in (request.headers.get("Accept", ""))

        if not username or not email or not password or password != confirm:
            error_message = "Проверьте введённые данные и повторите попытку."
            status_code = 400
            if wants_json:
                return jsonify({"detail": error_message}), status_code
            return render_template("register.html", error_message=error_message, status_message=status_message), status_code

        if query_db("SELECT 1 FROM users WHERE username=? OR email=?", (username, email), one=True):
            error_message = "Такой пользователь уже есть"
            status_code = 409
            if wants_json:
                return jsonify({"detail": error_message}), status_code
            return render_template("register.html", error_message=error_message, status_message=status_message), status_code

        salt = make_salt()
        pwd_hash = hash_password(password, salt)
        token = generate_token()
        now = datetime.datetime.utcnow().isoformat()
        query_db("INSERT INTO users (username, email, password_hash, password_salt, role, is_verified, verify_token, created_at, updated_at) VALUES (?, ?, ?, ?, 'viewer', 0, ?, ?, ?)",
                 (username, email, pwd_hash, salt, token, now, now))
        verify_link = f"{request.host_url.rstrip('/')}{url_for('verify')}?token={token}"

        try:
            send_email(email, "Подтверждение регистрации", f"Подтвердите ваш e-mail: {verify_link}")
        except RuntimeError:
            query_db("DELETE FROM users WHERE verify_token=?", (token,))
            error_message = "Не удалось отправить письмо. Проверьте настройки SMTP."
            status_code = 500
            if wants_json:
                return jsonify({"detail": error_message}), status_code
            return render_template("register.html", error_message=error_message, status_message=status_message), status_code

        status_message = f"Письмо отправлено на почту {email}. Проверьте входящие и папку спам."
        if wants_json:
            return jsonify({"detail": status_message}), 200
        return render_template("register.html", status_message=status_message, error_message=error_message), 200

    return render_template("register.html", status_message=status_message, error_message=error_message)

@app.route("/verify")
def verify():
    token = request.args.get("token", "")
    row = query_db("SELECT id FROM users WHERE verify_token=?", (token,), one=True)
    if not row:
        return "Неверный токен", 400
    query_db("UPDATE users SET is_verified=1, verify_token=NULL, updated_at=? WHERE id=?", (datetime.datetime.utcnow().isoformat(), row["id"]))
    return "E-mail подтверждён!"


def now_iso():
    return datetime.datetime.utcnow().isoformat()


def sanitize_name(name):
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", (name or "")).strip()
    if not cleaned:
        return ""
    return cleaned[:MAX_FILENAME_LEN]


def table_user_from_request():
    session_user_id = session.get("user_id")
    legacy_cookie_present = bool(request.cookies.get("tables_user_id"))
    if not session_user_id:
        if request.path.startswith("/api/tables"):
            reason = "missing_session"
            if legacy_cookie_present:
                reason = "missing_session_legacy_cookie_ignored"
            app.logger.info("[tables_auth] resolved_user_id=None username=None email=None reason=%s", reason)
        return None

    user = query_db("SELECT id, username, email FROM users WHERE id=?", (session_user_id,), one=True)
    if user is not None and not isinstance(user, dict):
        user = dict(user)

    if user is None:
        if request.path.startswith("/api/tables"):
            app.logger.info("[tables_auth] resolved_user_id=None username=None email=None reason=session_user_missing")
        return None

    if request.path.startswith("/api/tables"):
        app.logger.info(
            "[tables_auth] resolved_user_id=%s username=%s email=%s reason=session",
            user["id"],
            user.get("username"),
            user.get("email"),
        )
    return user


def storage_for_table(user_id, table_id):
    return os.path.join(STORAGE_ROOT, "users", str(user_id), "tables", str(table_id))


def get_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def parse_excel_preview(xlsx_path, preview_limit=EXCEL_PREVIEW_LIMIT):
    if openpyxl is None:
        raise RuntimeError("openpyxl недоступен")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise RuntimeError("В Excel не найдено ни одного листа")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return [], [], 0, ws.title

    headers = ["" if x is None else str(x).strip() for x in header_row]
    preview_rows = []
    total_rows = 0

    for row in rows_iter:
        total_rows += 1
        values = ["" if x is None else str(x).strip() for x in row]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        preview_rows.append(values)
        if len(preview_rows) >= preview_limit:
            break

    for _ in rows_iter:
        total_rows += 1

    return headers, preview_rows, total_rows, ws.title


def normalize_mapping(mapping):
    if not isinstance(mapping, dict):
        return {}
    normalized = {}
    used_cols = set()
    for field in MAPPING_FIELDS.keys():
        val = mapping.get(field)
        indexes = []
        if isinstance(val, int):
            indexes = [val]
        elif isinstance(val, list):
            indexes = [idx for idx in val if isinstance(idx, int)]

        for idx in indexes:
            if idx < 0 or idx in used_cols:
                continue
            normalized.setdefault(field, []).append(idx)
            used_cols.add(idx)

    for field, indexes in list(normalized.items()):
        if len(indexes) == 1 and field not in GROUPED_HEADER_RULES:
            normalized[field] = indexes[0]
        elif not indexes:
            normalized.pop(field, None)
    return normalized


def mapped_field_indexes(mapping, field):
    value = (mapping or {}).get(field)
    if isinstance(value, int):
        return [value]
    if isinstance(value, list):
        return [idx for idx in value if isinstance(idx, int) and idx >= 0]
    return []


def normalize_header_text(value):
    text = str(value or "")
    text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def mapping_signature(headers):
    normalized = [normalize_header_text(h) for h in (headers or [])]
    return hashlib.sha256("|".join(normalized).encode("utf-8")).hexdigest()


def apply_mapping_templates_and_presets(user_id, headers):
    normalized_headers = [normalize_header_text(h) for h in headers]
    signature = mapping_signature(headers)

    template = query_db(
        "SELECT mapping_json FROM mapping_templates WHERE user_id=? AND signature=?",
        (user_id, signature),
        one=True,
    )
    if template:
        mapping = normalize_mapping(json.loads(template["mapping_json"] or "{}"))
        return mapping, signature, "template"

    mapping = {}
    conflicts = []

    for idx, header in enumerate(normalized_headers):
        matches = []
        for label, aliases in HEADER_ALIASES.items():
            if label not in MAPPING_FIELDS:
                continue
            matched = False
            for alias in aliases:
                if header == normalize_header_text(alias):
                    matched = True
                    break
            if not matched and parse_grouped_header(label, header):
                matched = True
            if matched:
                matches.append(label)
        uniq = list(dict.fromkeys(matches))
        if len(uniq) == 1:
            label = uniq[0]
            mapping.setdefault(label, []).append(idx)
        elif len(uniq) > 1:
            conflicts.append({"idx": idx, "header": headers[idx], "labels": uniq})

    if conflicts:
        app.logger.warning("[table_tags] mapping conflicts detected: %s", conflicts)

    return normalize_mapping(mapping), signature, "tags_config"


def parse_grouped_header(field_name, normalized_header):
    prefixes = set()
    for prefix in GROUPED_HEADER_RULES.get(field_name) or []:
        prefix_norm = normalize_header_text(prefix)
        if prefix_norm:
            prefixes.add(prefix_norm)

    for alias in HEADER_ALIASES.get(field_name) or []:
        alias_norm = normalize_header_text(alias)
        if "/" in alias_norm:
            left = alias_norm.split("/", 1)[0].strip()
            if left:
                prefixes.add(left)

    if "/" not in normalized_header:
        return ""

    left, right = normalized_header.split("/", 1)
    left = left.strip()
    right = right.strip()
    if not right:
        return ""

    for prefix_norm in prefixes:
        if left == prefix_norm or left.startswith(prefix_norm + " "):
            return right
    return ""


def resolve_grouped_value(row_values, grouped_mappings):
    options = []
    for item in grouped_mappings:
        idx = item.get("idx")
        if idx is None or idx >= len(row_values):
            continue
        value = str(row_values[idx] or "").strip()
        if value:
            options.append(item.get("choice") or "")
    uniq = [x for x in dict.fromkeys(options) if x]
    if len(uniq) == 1:
        return uniq[0], uniq, False
    if len(uniq) > 1:
        return "", uniq, True
    return "", [], False


def detect_grouped_headers(headers, mapping=None):
    result = {k: [] for k in GROUPED_HEADER_RULES.keys()}
    seen = {k: set() for k in GROUPED_HEADER_RULES.keys()}

    for field_name in GROUPED_HEADER_RULES.keys():
        for idx in mapped_field_indexes(mapping, field_name):
            if idx >= len(headers or []):
                continue
            normalized = normalize_header_text(headers[idx])
            choice = parse_grouped_header(field_name, normalized) or normalized
            result[field_name].append({"idx": idx, "choice": choice, "header": headers[idx]})
            seen[field_name].add(idx)

    for idx, header in enumerate(headers or []):
        normalized = normalize_header_text(header)
        for field_name in GROUPED_HEADER_RULES.keys():
            if idx in seen[field_name]:
                continue
            choice = parse_grouped_header(field_name, normalized)
            if choice:
                result[field_name].append({"idx": idx, "choice": choice, "header": header})
                seen[field_name].add(idx)
    return result


def tag_type_for_key(tag_key):
    if tag_key in TAG_TYPE_OVERRIDES:
        return TAG_TYPE_OVERRIDES[tag_key]
    if tag_key in GROUPED_HEADER_RULES:
        return "grouped_choice"
    return "text"


def parse_entry_row_data(entry):
    raw = entry.get("row_data_json")
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    return {str(k): str(v or "").strip() for k, v in data.items()}


def get_entry_tag_values(entry, mapping, field_name, row_data=None):
    values = []
    seen = set()

    def add(raw):
        for part in split_multi_value(raw):
            if part in seen:
                continue
            seen.add(part)
            values.append(part)

    # Для тегов, которые имеют отдельные актуальные колонки в table_entries.
    add(entry.get(field_name))

    # Данные из новой схемы читаем из row_data_json по mapping.
    row_data = row_data if isinstance(row_data, dict) else parse_entry_row_data(entry)
    for idx in mapped_field_indexes(mapping, field_name):
        add(row_data.get(str(idx), ""))

    # Базовые поля программы: поддержка через актуальные поля table_entries.
    if field_name == "participant_fio":
        add(entry.get("fio"))
    elif field_name == "studio_name":
        add(entry.get("team"))
    elif field_name == "number_title":
        add(entry.get("number_title"))

    return values


def build_visible_columns(table_id, mapping):
    visible = []
    rows = query_db(
        """
        SELECT te.*
        FROM table_program_items tpi
        JOIN table_entries te ON te.id = tpi.entry_id AND te.table_id = tpi.table_id
        WHERE tpi.table_id=? AND tpi.kind='entry'
        ORDER BY tpi.sort_index, tpi.id
        """,
        (table_id,),
    )

    for field in MAPPING_FIELDS.keys():
        if field in mapping:
            visible.append(field)
            continue

        has_value = False
        for row in rows:
            entry = dict(row)
            row_data = parse_entry_row_data(entry)
            if get_entry_tag_values(entry, mapping, field, row_data=row_data):
                has_value = True
                break
            resolved = json.loads(entry.get("resolved_fields_json") or "{}")
            conflicts = json.loads(entry.get("conflicts_json") or "{}")
            if str(resolved.get(field) or "").strip() or any(str(x or "").strip() for x in (conflicts.get(field) or [])):
                has_value = True
                break

        if has_value:
            visible.append(field)
    return visible


def split_multi_value(raw_value):
    if raw_value is None:
        return []
    text = str(raw_value).strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;,\n]", text) if part and part.strip()]


def looks_like_url(value):
    text = str(value or "").strip().lower()
    return text.startswith("http://") or text.startswith("https://")


def entry_field_local_key(field_name):
    if field_name == "audio_url":
        return "audio_local"
    if field_name == "receipt_url":
        return "receipt_local"
    if field_name == "consent_url":
        return "consent_local"
    if field_name == "application_file":
        return "application_local"
    if field_name == "generic_file":
        return "generic_local"
    if field_name == "presentation_url":
        return "presentation_local"
    if field_name in {"work_photo", "work_sketch"}:
        return "generic_local"
    return ""


def build_tag_cell(table_id, user_id, entry, mapping, field_name, resolved_fields, conflicts):
    tag_type = tag_type_for_key(field_name)
    label = MAPPING_FIELDS.get(field_name, field_name)
    row_data = parse_entry_row_data(entry)
    values = get_entry_tag_values(entry, mapping, field_name, row_data=row_data)

    cell = {
        "key": field_name,
        "label": label,
        "type": tag_type,
        "value": "",
        "values": [],
        "links": [],
        "files": [],
        "conflict": None,
    }

    if tag_type == "grouped_choice":
        selected = str((resolved_fields or {}).get(field_name) or "").strip()
        options = [str(x).strip() for x in ((conflicts or {}).get(field_name) or []) if str(x).strip()]
        if not selected and values:
            selected = values[0]
        if selected and selected not in options:
            options.insert(0, selected)
        if options:
            cell["values"] = options
        if selected:
            cell["value"] = selected
        if len(options) > 1:
            cell["conflict"] = {"options": options, "selected": selected}
        return cell

    if tag_type == "links":
        links = [v for v in values if looks_like_url(v)]
        cell["links"] = [{"url": url, "title": url} for url in links]
        cell["value"] = ", ".join(links)
        return cell

    local_key = entry_field_local_key(field_name)
    local_value = entry.get(local_key) if local_key else ""
    files = []
    if local_key and local_value:
        local_parts = parse_local_file_values(local_value)
        for part in local_parts:
            exists = has_local_entry_file(table_id, user_id, entry.get("id"), local_key, part, placeholder_for_audio=(field_name == "audio_url"))
            if not exists:
                continue
            route_kind = PROGRAM_DOWNLOAD_ROUTE_BY_TAG.get(field_name)
            if route_kind:
                download_url = f"/api/tables/{table_id}/program/download/{route_kind}/{entry.get('program_item_id')}?part={part}"
            else:
                download_url = ""
            files.append({"name": os.path.basename(part), "path": part, "download_url": download_url})

    if tag_type == "files":
        cell["files"] = files
        cell["values"] = [f["name"] for f in files]
        return cell

    if tag_type == "files_or_links":
        links = [v for v in values if looks_like_url(v)]
        non_links = [v for v in values if not looks_like_url(v)]
        cell["links"] = [{"url": url, "title": url} for url in links]
        if files:
            cell["files"] = files
        else:
            cell["files"] = [{"name": n, "path": "", "download_url": ""} for n in non_links]
        return cell

    if values:
        cell["value"] = values[0]
        cell["values"] = values
    return cell


def save_mapping_template(user_id, signature, mapping):
    ts = now_iso()
    query_db(
        """
        INSERT INTO mapping_templates (user_id, signature, mapping_json, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(user_id, signature)
        DO UPDATE SET mapping_json=excluded.mapping_json, updated_at=excluded.updated_at
        """,
        (user_id, signature, json.dumps(normalize_mapping(mapping), ensure_ascii=False), ts),
    )


def upsert_mapping_preset(user_id, label, pattern, match_type, priority=100):
    if not pattern or label not in MAPPING_FIELDS:
        return
    ts = now_iso()
    query_db(
        """
        INSERT INTO user_mapping_presets (user_id, label, pattern, match_type, priority, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, label, pattern, match_type)
        DO UPDATE SET priority=excluded.priority, updated_at=excluded.updated_at
        """,
        (user_id, label, pattern, match_type, priority, ts),
    )


def update_table_yandex_status(table_id, status, error_text=None):
    query_db(
        "UPDATE table_workspaces SET yandex_status=?, yandex_last_error=?, yandex_last_checked_at=?, updated_at=? WHERE id=?",
        (status, (error_text or "")[:300] or None, now_iso(), now_iso(), table_id),
    )


def table_required_mapping_status(mapping):
    assigned_files = [f for f in FILE_MAPPING_FIELDS if f in mapping]
    if not assigned_files:
        return False, "Нужно назначить хотя бы один файловый/документный столбец: фонограмма/презентация/согласие"

    if any(f in mapping for f in ["audio_url", "presentation_url", "video_url", "application_file", "consent_url"]):
        if "number_title" not in mapping:
            return False, "Не назначена обязательная колонка: Навание номреа"
    return True, "ok"


def excel_path_for_table(user_id, table_id):
    base = storage_for_table(user_id, table_id)
    excel_path = os.path.join(base, "meta", "original.xlsx")
    if not os.path.exists(excel_path):
        excel_path = os.path.join(base, "meta", "excel_original.xlsx")
    return excel_path


def get_table_owner_id(table_id):
    row = query_db("SELECT user_id FROM table_workspaces WHERE id=?", (table_id,), one=True)
    if not row:
        return None
    return int(row["user_id"])


def collect_table_entries_stats(table_id):
    row = query_db(
        """
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN TRIM(COALESCE(number_title, '')) <> '' THEN 1 ELSE 0 END) AS with_number_title,
            SUM(CASE WHEN TRIM(COALESCE(fio, '')) <> '' THEN 1 ELSE 0 END) AS with_fio,
            SUM(CASE WHEN TRIM(COALESCE(audio_local, '')) <> '' THEN 1 ELSE 0 END) AS with_audio_local,
            SUM(CASE WHEN TRIM(COALESCE(receipt_local, '')) <> '' THEN 1 ELSE 0 END) AS with_receipt_local,
            SUM(CASE WHEN TRIM(COALESCE(presentation_local, '')) <> '' THEN 1 ELSE 0 END) AS with_presentation_local
        FROM table_entries
        WHERE table_id=?
        """,
        (table_id,),
        one=True,
    )
    return {k: int((row[k] if row else 0) or 0) for k in [
        "total",
        "with_number_title",
        "with_fio",
        "with_audio_local",
        "with_receipt_local",
        "with_presentation_local",
    ]}


def log_table_entries_stats(stage, table_id, *, previous=None):
    stats = collect_table_entries_stats(table_id)
    drop_msg = ""
    if previous:
        drops = []
        for key in ["with_audio_local", "with_receipt_local", "with_presentation_local"]:
            prev_val = int(previous.get(key) or 0)
            curr_val = int(stats.get(key) or 0)
            if prev_val > 0 and curr_val == 0:
                drops.append(f"{key}:{prev_val}->0")
        if drops:
            drop_msg = " drops=" + ",".join(drops)
    app.logger.info(
        "[tables_entries_stats] stage=%s table_id=%s total=%s with_number_title=%s with_fio=%s with_audio_local=%s with_receipt_local=%s with_presentation_local=%s%s",
        stage,
        table_id,
        stats["total"],
        stats["with_number_title"],
        stats["with_fio"],
        stats["with_audio_local"],
        stats["with_receipt_local"],
        stats["with_presentation_local"],
        drop_msg,
    )
    return stats


def _name_token(value):
    return re.sub(r"[^a-zа-я0-9]+", "", (value or "").lower(), flags=re.IGNORECASE)


def repair_table_local_paths(table_id, user_id):
    base = storage_for_table(user_id, table_id)
    field_config = {
        "audio_local": ("phonograms", "phonogram"),
        "receipt_local": ("receipts", "receipt"),
        "presentation_local": ("presentations", "presentation"),
    }
    files_by_field = {}
    for field_name, (folder, _) in field_config.items():
        folder_path = Path(base) / folder
        files_by_field[field_name] = [
            p for p in folder_path.iterdir()
            if p.is_file()
        ] if folder_path.exists() else []

    rows = query_db(
        "SELECT id, row_id, number_title, fio, audio_local, receipt_local, presentation_local FROM table_entries WHERE table_id=? ORDER BY id",
        (table_id,),
    )
    repaired = {"audio_local": 0, "receipt_local": 0, "presentation_local": 0}
    used_paths = set()
    for row in rows:
        row = dict(row)
        updates = {}
        for field_name, (folder, fallback_prefix) in field_config.items():
            current_value = (row.get(field_name) or "").strip()
            _, current_full_path = resolve_table_local_path(table_id, user_id, current_value)
            if current_value and os.path.exists(current_full_path):
                used_paths.add(Path(current_full_path).resolve(strict=False))
                continue

            row_id = int(row.get("row_id") or 0)
            expected_base = make_safe_basename(row.get("number_title"), row.get("fio"), fallback=f"{fallback_prefix}-{row_id}")
            expected_token = _name_token(expected_base)
            fio_token = _name_token(row.get("fio"))
            title_token = _name_token(row.get("number_title"))
            best_score = -1
            best_file = None
            for file_path in files_by_field[field_name]:
                candidate_resolved = file_path.resolve(strict=False)
                if candidate_resolved in used_paths:
                    continue
                stem_token = _name_token(file_path.stem)
                score = 0
                if expected_token and stem_token == expected_token:
                    score += 5
                elif expected_token and (expected_token in stem_token or stem_token in expected_token):
                    score += 3
                if row_id and str(row_id) in file_path.stem:
                    score += 2
                if fio_token and fio_token in stem_token:
                    score += 1
                if title_token and title_token in stem_token:
                    score += 1
                if score > best_score:
                    best_score = score
                    best_file = file_path

            if best_file is not None and best_score > 0:
                rel = f"{folder}/{best_file.name}"
                updates[field_name] = rel
                used_paths.add(best_file.resolve(strict=False))
                repaired[field_name] += 1

        if updates:
            query_db(
                "UPDATE table_entries SET audio_local=COALESCE(?, audio_local), receipt_local=COALESCE(?, receipt_local), presentation_local=COALESCE(?, presentation_local) WHERE id=?",
                (updates.get("audio_local"), updates.get("receipt_local"), updates.get("presentation_local"), row["id"]),
            )

    app.logger.info(
        "[tables_entries_repair] table_id=%s repaired_audio=%s repaired_receipt=%s repaired_presentation=%s",
        table_id,
        repaired["audio_local"],
        repaired["receipt_local"],
        repaired["presentation_local"],
    )
    return repaired


def rebuild_entries_from_excel(table_id, user_id, mapping, *, preserve_files=True, clear_existing=True):
    excel_path = excel_path_for_table(user_id, table_id)
    if not os.path.exists(excel_path):
        raise RuntimeError("Excel файл не найден")
    if openpyxl is None:
        raise RuntimeError("openpyxl недоступен")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.values)
    data_rows = rows[1:] if rows else []

    existing_by_row_id = {}
    if preserve_files:
        existing_rows = query_db(
            "SELECT row_id, audio_local, receipt_local, presentation_local FROM table_entries WHERE table_id=?",
            (table_id,),
        )
        existing_by_row_id = {
            int(r["row_id"]): dict(r)
            for r in existing_rows
            if r["row_id"] is not None
        }

    if clear_existing:
        query_db("DELETE FROM table_entries WHERE table_id=?", (table_id,))

    stats = {
        "total": 0,
        "with_number_title": 0,
        "with_fio": 0,
        "with_audio_local": 0,
        "with_receipt_local": 0,
        "with_presentation_local": 0,
    }

    grouped_headers = detect_grouped_headers(rows[0] if rows else [], mapping)

    for i, row in enumerate(data_rows, start=1):
        row_id = i + 1
        row_values = ["" if x is None else str(x).strip() for x in row]
        row_data = {str(idx): val for idx, val in enumerate(row_values)}

        resolved_fields = {}
        conflicts = {}
        for grouped_field, grouped_map in grouped_headers.items():
            value, options, conflicted = resolve_grouped_value(row_values, grouped_map)
            if conflicted:
                conflicts[grouped_field] = options
            if value:
                resolved_fields[grouped_field] = value

        number_title = get_cell(row_values, mapping.get("number_title"))
        fio = get_cell(row_values, mapping.get("participant_fio"))
        team = get_cell(row_values, mapping.get("studio_name"))
        unique_key = f"{row_id}|{number_title}|{fio}"

        prev = existing_by_row_id.get(row_id) or {}
        audio_local = (prev.get("audio_local") or "") if preserve_files else ""
        receipt_local = (prev.get("receipt_local") or "") if preserve_files else ""
        presentation_local = (prev.get("presentation_local") or "") if preserve_files else ""

        query_db(
            """
            INSERT INTO table_entries (table_id, row_id, number_title, fio, team, unique_key, audio_url, receipt_url, consent_url, application_file, generic_file, video_url, presentation_url, audio_local, receipt_local, consent_local, application_local, generic_local, presentation_local, row_data_json, resolved_fields_json, conflicts_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                table_id,
                row_id,
                number_title,
                fio,
                team,
                unique_key,
                get_cell(row_values, mapping.get("audio_url")),
                get_cell(row_values, mapping.get("receipt_url")),
                get_cell(row_values, mapping.get("consent_url")),
                get_cell(row_values, mapping.get("application_file")),
                get_cell(row_values, mapping.get("generic_file")),
                get_cell(row_values, mapping.get("video_url")),
                get_cell(row_values, mapping.get("presentation_url")),
                audio_local,
                receipt_local,
                "",
                "",
                "",
                presentation_local,
                json.dumps(row_data, ensure_ascii=False),
                json.dumps(resolved_fields, ensure_ascii=False),
                json.dumps(conflicts, ensure_ascii=False),
                now_iso(),
            ),
        )

        stats["total"] += 1
        if number_title:
            stats["with_number_title"] += 1
        if fio:
            stats["with_fio"] += 1
        if audio_local:
            stats["with_audio_local"] += 1
        if receipt_local:
            stats["with_receipt_local"] += 1
        if presentation_local:
            stats["with_presentation_local"] += 1

    if preserve_files:
        repair_table_local_paths(table_id, user_id)
        stats = collect_table_entries_stats(table_id)
    return stats


def make_safe_basename(*parts, fallback="row"):
    raw = " - ".join([str(x).strip() for x in parts if str(x).strip()])
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", raw).strip(" .")
    if not cleaned:
        cleaned = fallback
    return cleaned[:MAX_FILENAME_LEN]


def add_row_suffix_if_needed(base_name, row_id, used_names):
    candidate = base_name
    if candidate.lower() in used_names:
        candidate = f"{base_name}-{row_id}"[:MAX_FILENAME_LEN]
    used_names.add(candidate.lower())
    return candidate


class YandexAuthRequiredError(RuntimeError):
    pass


class DownloadValidationError(RuntimeError):
    pass


def extension_from_url(url, fallback="bin"):
    name = url.split("?")[0].rstrip("/").split("/")[-1]
    if "." in name:
        ext = name.rsplit(".", 1)[-1].lower()
        return ext[:10]
    return fallback


def _response_looks_like_html(response):
    ctype = (response.headers.get("Content-Type") or "").lower()
    cdisp = (response.headers.get("Content-Disposition") or "").lower()
    final_url = str(response.url or "")
    final_url_l = final_url.lower()
    body_prefix = (response.content or b"")[:512].lstrip().lower()

    if response.status_code in (401, 403) or "passport.yandex" in final_url_l:
        raise YandexAuthRequiredError("Нужен вход администратора в Яндекс")

    if "text/html" in ctype:
        return True, "ответ имеет Content-Type text/html"
    if ctype.startswith("text/") and "attachment" not in cdisp:
        return True, f"ответ имеет текстовый Content-Type: {ctype}"
    if "forms.yandex.ru/u/files" in final_url_l and "attachment" not in cdisp and "audio" not in ctype:
        return True, "URL файла Yandex Forms вернул страницу вместо бинарного файла"
    if body_prefix.startswith(b"<!doctype html") or body_prefix.startswith(b"<html"):
        return True, "первые байты ответа похожи на HTML"
    return False, ""


def download_with_retries(req_session, url, out_path):
    last_error = None
    for _ in range(DOWNLOAD_RETRIES):
        try:
            r = req_session.get(url, timeout=30, allow_redirects=True)
            r.raise_for_status()
            is_html, reason = _response_looks_like_html(r)
            if is_html:
                raise DownloadValidationError(
                    f"вместо файла получен HTML/текст ({reason}); status={r.status_code}; "
                    f"content_type={r.headers.get('Content-Type')}; "
                    f"content_disposition={r.headers.get('Content-Disposition')}; final_url={r.url}"
                )
            with open(out_path, "wb") as f:
                f.write(r.content)
            return
        except YandexAuthRequiredError:
            raise
        except Exception as e:
            last_error = str(e)
            time.sleep(1)
    raise RuntimeError(last_error or "download failed")


def parse_local_file_values(local_value):
    text = str(local_value or "").strip()
    if not text:
        return []
    if text.startswith("["):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x or "").strip()]
        except Exception:
            pass
    return [text]


def short_error_message(exc, fallback="Ошибка обработки"):
    msg = str(exc or "").strip() or fallback
    return msg[:300]


def process_table_download(table_id, user_id):
    t = query_db("SELECT * FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user_id), one=True)
    if not t:
        return

    try:
        mapping = normalize_mapping(json.loads(t["mapping_json"] or "{}"))
        ok_to_run, reason = table_required_mapping_status(mapping)
        if not ok_to_run:
            query_db(
                "UPDATE table_workspaces SET status='error', total_count=0, processed_count=0, download_cursor_row_id=NULL, progress=0, last_error=?, updated_at=? WHERE id=?",
                (reason, now_iso(), table_id),
            )
            return

        owner_user_id = get_table_owner_id(table_id)
        if owner_user_id is None:
            query_db(
                "UPDATE table_workspaces SET status='error', total_count=0, processed_count=0, download_cursor_row_id=NULL, progress=0, last_error=?, updated_at=? WHERE id=?",
                ("Таблица не найдена", now_iso(), table_id),
            )
            return
        if owner_user_id != int(user_id):
            app.logger.warning(
                "[tables_storage_owner_mismatch] table_id=%s worker_user_id=%s owner_user_id=%s",
                table_id,
                user_id,
                owner_user_id,
            )
        user_id = owner_user_id

        base = storage_for_table(user_id, table_id)
        excel_path = excel_path_for_table(user_id, table_id)
        if not os.path.exists(excel_path) or openpyxl is None:
            query_db(
                "UPDATE table_workspaces SET status='error', total_count=0, processed_count=0, download_cursor_row_id=NULL, progress=0, last_error=?, updated_at=? WHERE id=?",
                ("Excel файл не найден или openpyxl недоступен", now_iso(), table_id),
            )
            return

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            query_db(
                "UPDATE table_workspaces SET status='done', total_count=0, processed_count=0, download_cursor_row_id=NULL, progress=100, last_error=NULL, updated_at=? WHERE id=?",
                (now_iso(), table_id),
            )
            return

        headers = ["" if x is None else str(x).strip() for x in (rows[0] or [])]
        data_rows = rows[1:]
        total = len(data_rows)
        existing_processed = int(t["processed_count"] or 0)
        existing_cursor = t["download_cursor_row_id"]
        is_resume = str(t["status"] or "").lower() in {"auth_required", "paused", "downloading_partial"}

        query_db(
            "UPDATE table_workspaces SET status='downloading', total_count=?, processed_count=?, progress=?, last_error=NULL, updated_at=? WHERE id=?",
            (
                total,
                existing_processed if is_resume else 0,
                (100 if total == 0 else int(((existing_processed if is_resume else 0) / total) * 100)),
                now_iso(),
                table_id,
            ),
        )

        try:
            cookies = read_yandex_cookies_from_chromium_profile()
        except Exception as exc:
            error_text = short_error_message(exc, "Нужен вход администратора в Яндекс.")
            update_table_yandex_status(table_id, "auth_required", error_text)
            query_db(
                "UPDATE table_workspaces SET status='auth_required', last_error=?, updated_at=? WHERE id=?",
                (error_text, now_iso(), table_id),
            )
            return

        access_ok, final_url = check_yandex_auth(cookies)
        if not access_ok:
            error_text = "Нужен вход администратора в Яндекс."
            update_table_yandex_status(table_id, "auth_required", f"Нужен вход администратора в Яндекс: {final_url}")
            query_db(
                "UPDATE table_workspaces SET status='auth_required', last_error=?, updated_at=? WHERE id=?",
                (error_text, now_iso(), table_id),
            )
            return
        update_table_yandex_status(table_id, "connected", None)

        req_session = requests.Session()
        apply_cookies_to_requests_session(req_session, cookies)

        def persist_auth_required_stop(row_id, reason):
            query_db(
                "UPDATE table_workspaces SET status='auth_required', download_cursor_row_id=?, last_error=?, updated_at=? WHERE id=?",
                (row_id, reason, now_iso(), table_id),
            )
            app.logger.warning(
                "[tables_download] table_id=%s stopped_auth_required row_id=%s processed=%s/%s reason=%s",
                table_id,
                row_id,
                int(query_db("SELECT processed_count FROM table_workspaces WHERE id=?", (table_id,), one=True)["processed_count"] or 0),
                total,
                reason,
            )

        def handle_auth_required_with_refresh(row_id):
            app.logger.info("[tables_download] table_id=%s auto_refresh_on_auth_required row_id=%s", table_id, row_id)
            result = try_refresh_yandex_session(table_id, user_id, reason="download_auto")
            if result.get("status") != "ok":
                reason = "Нужен вход администратора в Яндекс."
                update_table_yandex_status(table_id, "auth_required", reason)
                persist_auth_required_stop(row_id, reason)
                return False

            try:
                refreshed = read_yandex_cookies_from_chromium_profile()
            except Exception:
                reason = "Нужен вход администратора в Яндекс."
                persist_auth_required_stop(row_id, reason)
                return False
            req_session.cookies.clear()
            apply_cookies_to_requests_session(req_session, refreshed)
            return True

        for folder in ["phonograms", "receipts", "presentations", "meta"]:
            os.makedirs(os.path.join(base, folder), exist_ok=True)

        existing_rows = query_db(
            "SELECT id, row_id, audio_local, receipt_local, presentation_local FROM table_entries WHERE table_id=?",
            (table_id,),
        )
        existing_by_row_id = {
            int(r["row_id"]): dict(r)
            for r in existing_rows
            if r["row_id"] is not None
        }

        before_download_stats = collect_table_entries_stats(table_id)
        if not is_resume:
            query_db("DELETE FROM table_entries WHERE table_id=?", (table_id,))
            existing_by_row_id = {}

        used_names = {"phonograms": set(), "receipts": set(), "presentations": set()}
        processed_count = existing_processed if is_resume else 0
        skipped_completed_rows = 0
        downloaded_files = 0
        start_row_id = int(existing_cursor or 2) if is_resume else 2
        app.logger.info(
            "[tables_download] table_id=%s mode=%s resume_from_row=%s processed=%s/%s",
            table_id,
            "resume" if is_resume else "start",
            start_row_id,
            processed_count,
            total,
        )

        grouped_headers = detect_grouped_headers(headers, mapping)
        for i, row in enumerate(data_rows, start=1):
            row_id = i + 1
            if row_id < start_row_id:
                continue

            row_values = ["" if x is None else str(x).strip() for x in row]
            row_data = {str(idx): val for idx, val in enumerate(row_values)}
            resolved_fields = {}
            conflicts = {}
            for grouped_field, grouped_map in grouped_headers.items():
                value, options, conflicted = resolve_grouped_value(row_values, grouped_map)
                if conflicted:
                    conflicts[grouped_field] = options
                if value:
                    resolved_fields[grouped_field] = value

            number_title = get_cell(row_values, mapping.get("number_title"))
            fio = get_cell(row_values, mapping.get("participant_fio"))
            team = get_cell(row_values, mapping.get("studio_name"))
            unique_key = f"{row_id}|{number_title}|{fio}"

            prev = existing_by_row_id.get(row_id) or {}
            prev_audio_local = (prev.get("audio_local") or "").strip()
            prev_receipt_local = (prev.get("receipt_local") or "").strip()
            prev_presentation_local = (prev.get("presentation_local") or "").strip()
            entry_id = prev.get("id")

            audio_url = get_cell(row_values, mapping.get("audio_url"))
            receipt_url = get_cell(row_values, mapping.get("receipt_url"))
            presentation_url = get_cell(row_values, mapping.get("presentation_url"))

            audio_done = bool(prev_audio_local and os.path.exists(os.path.join(base, prev_audio_local))) if audio_url else True
            receipt_done = bool(prev_receipt_local and os.path.exists(os.path.join(base, prev_receipt_local))) if receipt_url else True
            presentation_done = bool(prev_presentation_local and os.path.exists(os.path.join(base, prev_presentation_local))) if presentation_url else True

            row_fully_done = audio_done and receipt_done and presentation_done
            if row_fully_done and entry_id:
                skipped_completed_rows += 1
                processed_count += 1
                progress = 100 if total == 0 else int((processed_count / total) * 100)
                query_db(
                    "UPDATE table_workspaces SET processed_count=?, progress=?, updated_at=? WHERE id=?",
                    (processed_count, progress, now_iso(), table_id),
                )
                continue

            if entry_id:
                query_db(
                    """
                    UPDATE table_entries
                    SET number_title=?, fio=?, team=?, unique_key=?, audio_url=?, receipt_url=?, consent_url=?, application_file=?, generic_file=?, video_url=?, presentation_url=?, row_data_json=?, resolved_fields_json=?, conflicts_json=?
                    WHERE id=?
                    """,
                    (
                        number_title,
                        fio,
                        team,
                        unique_key,
                        audio_url,
                        receipt_url,
                        get_cell(row_values, mapping.get("consent_url")),
                        get_cell(row_values, mapping.get("application_file")),
                        get_cell(row_values, mapping.get("generic_file")),
                        get_cell(row_values, mapping.get("video_url")),
                        presentation_url,
                        json.dumps(row_data, ensure_ascii=False),
                        json.dumps(resolved_fields, ensure_ascii=False),
                        json.dumps(conflicts, ensure_ascii=False),
                        entry_id,
                    ),
                )
            else:
                query_db(
                    """
                    INSERT INTO table_entries (table_id, row_id, number_title, fio, team, unique_key, audio_url, receipt_url, consent_url, application_file, generic_file, video_url, presentation_url, audio_local, receipt_local, consent_local, application_local, generic_local, presentation_local, row_data_json, resolved_fields_json, conflicts_json, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        table_id,
                        row_id,
                        number_title,
                        fio,
                        team,
                        unique_key,
                        audio_url,
                        receipt_url,
                        get_cell(row_values, mapping.get("consent_url")),
                        get_cell(row_values, mapping.get("application_file")),
                        get_cell(row_values, mapping.get("generic_file")),
                        get_cell(row_values, mapping.get("video_url")),
                        presentation_url,
                        prev_audio_local,
                        prev_receipt_local,
                        "",
                        "",
                        "",
                        prev_presentation_local,
                        json.dumps(row_data, ensure_ascii=False),
                        json.dumps(resolved_fields, ensure_ascii=False),
                        json.dumps(conflicts, ensure_ascii=False),
                        now_iso(),
                    ),
                )
                entry_id = query_db("SELECT last_insert_rowid() AS id", one=True)["id"]

            query_db(
                "UPDATE table_workspaces SET download_cursor_row_id=?, updated_at=? WHERE id=?",
                (row_id, now_iso(), table_id),
            )

            audio_local = prev_audio_local
            receipt_local = prev_receipt_local
            presentation_local = prev_presentation_local

            if "audio_url" in mapping and audio_url and not audio_done:
                phonogram_base = add_row_suffix_if_needed(
                    make_safe_basename(number_title, fio, fallback=f"phonogram-{row_id}"),
                    row_id,
                    used_names["phonograms"],
                )
                ext = extension_from_url(audio_url, "mp3")
                audio_name = f"{phonogram_base}.{ext}"
                audio_path = os.path.join(base, "phonograms", audio_name)
                try:
                    download_with_retries(req_session, audio_url, audio_path)
                    audio_local = os.path.join("phonograms", audio_name)
                    downloaded_files += 1
                except YandexAuthRequiredError:
                    if not handle_auth_required_with_refresh(row_id):
                        return
                    try:
                        download_with_retries(req_session, audio_url, audio_path)
                        audio_local = os.path.join("phonograms", audio_name)
                        downloaded_files += 1
                    except YandexAuthRequiredError:
                        persist_auth_required_stop(row_id, "Нужен вход администратора в Яндекс.")
                        return
                except Exception as exc:
                    app.logger.warning(
                        "[tables_download_audio_invalid] table_id=%s row_id=%s url=%s reason=%s",
                        table_id,
                        row_id,
                        audio_url,
                        short_error_message(exc, fallback="Ошибка скачивания фонограммы"),
                    )

            if receipt_url and not receipt_done:
                payer = get_cell(row_values, mapping.get("receipt_payer"))
                receipt_base = add_row_suffix_if_needed(
                    make_safe_basename(payer, fallback=f"receipt-{row_id}"),
                    row_id,
                    used_names["receipts"],
                )
                ext = extension_from_url(receipt_url, "pdf")
                receipt_name = f"{receipt_base}.{ext}"
                receipt_path = os.path.join(base, "receipts", receipt_name)
                try:
                    download_with_retries(req_session, receipt_url, receipt_path)
                    receipt_local = os.path.join("receipts", receipt_name)
                    downloaded_files += 1
                except YandexAuthRequiredError:
                    if not handle_auth_required_with_refresh(row_id):
                        return
                    try:
                        download_with_retries(req_session, receipt_url, receipt_path)
                        receipt_local = os.path.join("receipts", receipt_name)
                        downloaded_files += 1
                    except YandexAuthRequiredError:
                        persist_auth_required_stop(row_id, "Нужен вход администратора в Яндекс.")
                        return
                except Exception:
                    pass

            if presentation_url and not presentation_done:
                presentation_base = add_row_suffix_if_needed(
                    make_safe_basename(number_title, fio, fallback=f"presentation-{row_id}"),
                    row_id,
                    used_names["presentations"],
                )
                ext = extension_from_url(presentation_url, "bin")
                presentation_name = f"{presentation_base}.{ext}"
                presentation_path = os.path.join(base, "presentations", presentation_name)
                try:
                    download_with_retries(req_session, presentation_url, presentation_path)
                    presentation_local = os.path.join("presentations", presentation_name)
                    downloaded_files += 1
                except YandexAuthRequiredError:
                    if not handle_auth_required_with_refresh(row_id):
                        return
                    try:
                        download_with_retries(req_session, presentation_url, presentation_path)
                        presentation_local = os.path.join("presentations", presentation_name)
                        downloaded_files += 1
                    except YandexAuthRequiredError:
                        persist_auth_required_stop(row_id, "Нужен вход администратора в Яндекс.")
                        return
                except Exception:
                    pass

            query_db(
                "UPDATE table_entries SET audio_local=?, receipt_local=?, presentation_local=? WHERE id=?",
                (audio_local, receipt_local, presentation_local, entry_id),
            )

            processed_count += 1
            progress = 100 if total == 0 else int((processed_count / total) * 100)
            query_db(
                "UPDATE table_workspaces SET processed_count=?, progress=?, updated_at=? WHERE id=?",
                (processed_count, progress, now_iso(), table_id),
            )

        repair_table_local_paths(table_id, user_id)
        log_table_entries_stats("download_done", table_id, previous=before_download_stats)

        app.logger.info(
            "[tables_download] table_id=%s completed processed=%s/%s skipped_completed_rows=%s downloaded_files=%s",
            table_id,
            processed_count,
            total,
            skipped_completed_rows,
            downloaded_files,
        )
        query_db(
            "UPDATE table_workspaces SET status='done', processed_count=?, download_cursor_row_id=NULL, progress=100, last_error=NULL, updated_at=? WHERE id=?",
            (processed_count, now_iso(), table_id),
        )
    except Exception as exc:
        app.logger.exception("[tables_download] table_id=%s failed", table_id)
        query_db(
            "UPDATE table_workspaces SET status='error', last_error=?, updated_at=? WHERE id=?",
            (short_error_message(exc), now_iso(), table_id),
        )


def cleanup_old_tables():
    while True:
        cutoff = (datetime.datetime.utcnow() - datetime.timedelta(days=RETENTION_DAYS)).isoformat()
        old_tables = query_db("SELECT id, user_id FROM table_workspaces WHERE created_at < ?", (cutoff,))
        for t in old_tables:
            table_id, user_id = t["id"], t["user_id"]
            query_db("DELETE FROM table_entries WHERE table_id=?", (table_id,))
            query_db("DELETE FROM table_workspaces WHERE id=?", (table_id,))
            shutil.rmtree(storage_for_table(user_id, table_id), ignore_errors=True)
        time.sleep(24 * 3600)


_tables_cleanup_started = False


def repair_empty_entries_for_finalized_tables():
    rows = query_db(
        """
        SELECT id, user_id, mapping_json
        FROM table_workspaces
        WHERE is_finalized=1
        """
    )
    for t in rows:
        table_id = t["id"]
        user_id = int(t["user_id"])
        mapping = normalize_mapping(json.loads(t.get("mapping_json") or "{}"))
        before_stats = collect_table_entries_stats(table_id)
        total = before_stats["total"]
        with_number_title = before_stats["with_number_title"]
        with_fio = before_stats["with_fio"]
        try:
            if total == 0 or (with_number_title == 0 and with_fio == 0):
                rebuild_stats = rebuild_entries_from_excel(table_id, user_id, mapping, preserve_files=True, clear_existing=True)
                app.logger.info(
                    "[tables_entries_rebuild] stage=startup_repair table_id=%s total=%s with_number_title=%s with_fio=%s with_audio_local=%s with_receipt_local=%s with_presentation_local=%s",
                    table_id,
                    rebuild_stats["total"],
                    rebuild_stats["with_number_title"],
                    rebuild_stats["with_fio"],
                    rebuild_stats["with_audio_local"],
                    rebuild_stats["with_receipt_local"],
                    rebuild_stats["with_presentation_local"],
                )
            repair_table_local_paths(table_id, user_id)
            log_table_entries_stats("startup_repair", table_id, previous=before_stats)
        except Exception as exc:
            app.logger.warning("[tables_entries_rebuild] stage=startup_repair table_id=%s skipped error=%s", table_id, str(exc))


def start_tables_background_jobs():
    global _tables_cleanup_started
    if _tables_cleanup_started:
        return
    os.makedirs(STORAGE_ROOT, exist_ok=True)
    APP_DATA_ROOT.mkdir(parents=True, exist_ok=True)
    repair_empty_entries_for_finalized_tables()
    threading.Thread(target=cleanup_old_tables, daemon=True).start()
    _tables_cleanup_started = True


@app.route("/tables")
@login_required
def tables_page():
    return render_template("tables.html")


@app.route("/api/tables/send_code", methods=["POST"])
def tables_send_code():
    return jsonify({"detail": "Используйте основной вход /login"}), 410


@app.route("/api/tables/verify_code", methods=["POST"])
def tables_verify_code():
    return jsonify({"detail": "Используйте основной вход /login"}), 410


@app.route("/api/tables", methods=["GET"])
def list_tables():
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    rows = query_db(
        "SELECT id, title, status, total_count, processed_count, download_cursor_row_id, progress, last_error, created_at, mapping_json, yandex_status, yandex_last_error, yandex_last_checked_at, is_finalized, finalized_at, program_updated_at FROM table_workspaces WHERE user_id=? ORDER BY id DESC",
        (user["id"],),
    )
    result = []
    for r in rows:
        item = dict(r)
        item["yandex_status"] = item.get("yandex_status") or "disconnected"
        item["mapping"] = normalize_mapping(json.loads(item.get("mapping_json") or "{}"))
        result.append(item)
    return jsonify(result)


@app.route("/api/tables", methods=["POST"])
def create_table():
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    title = (request.form.get("title") or "").strip()
    if not title:
        return jsonify({"detail": "Название обязательно"}), 400
    cnt = query_db("SELECT COUNT(*) AS c FROM table_workspaces WHERE user_id=?", (user["id"],), one=True)["c"]
    if cnt >= MAX_TABLES_PER_USER:
        return jsonify({"detail": "Достигнут лимит 10 таблиц"}), 400
    ts = now_iso()
    query_db(
        "INSERT INTO table_workspaces (user_id, title, created_at, updated_at, yandex_status, yandex_last_checked_at) VALUES (?, ?, ?, ?, 'disconnected', ?)",
        (user["id"], title, ts, ts, ts),
    )
    return jsonify({"status": "ok"})


def apply_cookies_to_requests_session(req_session, cookies):
    for c in cookies:
        name = c.get("name")
        value = c.get("value")
        if not name or value is None:
            continue
        req_session.cookies.set(name, value, domain=c.get("domain") or ".yandex.ru", path=c.get("path") or "/")


def yandex_profile_cookie_file():
    return Path(YANDEX_PROFILE_DIR) / "Default" / "Cookies"


def read_yandex_cookies_from_chromium_profile():
    cookie_db = yandex_profile_cookie_file()
    if not cookie_db.exists():
        raise RuntimeError(f"Файл cookies Chromium не найден: {cookie_db}")

    tmp_copy = APP_DATA_ROOT / "tmp" / "yandex_cookies_read.sqlite"
    tmp_copy.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(cookie_db, tmp_copy)

    con = sqlite3.connect(str(tmp_copy))
    con.row_factory = sqlite3.Row
    try:
        rows = con.execute(
            """
            SELECT host_key, name, value, path
            FROM cookies
            WHERE host_key LIKE '%yandex.ru%' OR host_key LIKE '%yandex.net%'
            """
        ).fetchall()
    finally:
        con.close()
        tmp_copy.unlink(missing_ok=True)

    cookies = []
    for row in rows:
        cookies.append(
            {
                "domain": row["host_key"],
                "name": row["name"],
                "value": row["value"],
                "path": row["path"] or "/",
            }
        )

    if not cookies:
        raise RuntimeError("В профиле Chromium не найдены cookies Яндекса")
    return cookies


def yandex_cookie_domains(cookies):
    domains = set()
    for c in cookies:
        domain = (c.get("domain") or "").strip().lower()
        if domain:
            domains.add(domain.lstrip('.'))
    return sorted(domains)


def check_yandex_auth(cookies):
    req = requests.Session()
    apply_cookies_to_requests_session(req, cookies)
    resp = req.get("https://disk.yandex.ru/client/disk", timeout=15, allow_redirects=True)
    final_url = (resp.url or "").lower()
    if resp.status_code in (401, 403) or "passport.yandex" in final_url:
        return False, resp.url or ""

    if YANDEX_FORMS_TEST_URL:
        probe = req.get(YANDEX_FORMS_TEST_URL, timeout=20, allow_redirects=True)
        probe_url = (probe.url or "").lower()
        if probe.status_code in (401, 403) or "passport.yandex" in probe_url:
            return False, probe.url or ""
    return True, resp.url or ""


def has_global_yandex_session():
    try:
        cookies = read_yandex_cookies_from_chromium_profile()
        app.logger.info(
            "[yandex_connect] cookies_count=%s domains=%s",
            len(cookies),
            ",".join(yandex_cookie_domains(cookies)),
        )
        ok, final_url = check_yandex_auth(cookies)
        if not ok:
            app.logger.info("[yandex_connect] auth_required final_url=%s", final_url)
        return ok
    except Exception as exc:
        app.logger.info("[yandex_connect] read_failed error=%s", str(exc))
        return False


def table_belongs_to_user(table_id, user_id):
    return query_db("SELECT id FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user_id), one=True)


def run_shell_command(cmd, timeout=30):
    if not cmd:
        return False, ""
    try:
        proc = subprocess.run(
            cmd,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout,
            text=True,
        )
        output = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
        return proc.returncode == 0, output.strip()[:500]
    except Exception as exc:
        return False, str(exc)[:300]


def trigger_yandex_profile_network_refresh():
    refresh_url = YANDEX_REFRESH_URL or "https://disk.yandex.ru/client/disk"
    commands = []
    if YANDEX_CHROMIUM_RESTART_CMD:
        commands.append(YANDEX_CHROMIUM_RESTART_CMD.replace("{url}", refresh_url))
    if YANDEX_CHROMIUM_OPEN_CMD:
        commands.append(YANDEX_CHROMIUM_OPEN_CMD.replace("{url}", refresh_url))

    for cmd in commands:
        ok, out = run_shell_command(cmd, timeout=40)
        app.logger.info("[yandex_refresh] command=%s success=%s", cmd, ok)
        if ok:
            if out:
                app.logger.info("[yandex_refresh] command_output=%s", out)
            return True, None
        if out:
            app.logger.warning("[yandex_refresh] command_failed_output=%s", out)

    req = requests.Session()
    req.get(refresh_url, timeout=20, allow_redirects=True)
    return True, None


def persist_table_yandex_session(table_id, cookies):
    query_db(
        "UPDATE table_workspaces SET yandex_session_json=?, updated_at=? WHERE id=?",
        (json.dumps(cookies, ensure_ascii=False), now_iso(), table_id),
    )


def try_refresh_yandex_session(table_id, user_id, reason="manual"):
    if not table_belongs_to_user(table_id, user_id):
        return {"status": "missing", "detail": "Таблица не найдена"}

    app.logger.info("[yandex_refresh] table_id=%s reason=%s start", table_id, reason)
    try:
        trigger_yandex_profile_network_refresh()
    except Exception as exc:
        app.logger.warning("[yandex_refresh] table_id=%s trigger_failed error=%s", table_id, str(exc))

    try:
        cookies = read_yandex_cookies_from_chromium_profile()
    except Exception as exc:
        error_text = short_error_message(exc, "Нужен вход администратора в Яндекс")
        update_table_yandex_status(table_id, "auth_required", error_text)
        return {
            "status": "need_login",
            "detail": "Требуется вход администратора",
            "yandex_status": "auth_required",
            "cookies_count": 0,
        }

    app.logger.info("[yandex_refresh] table_id=%s reason=%s cookies_count=%s", table_id, reason, len(cookies))
    persist_table_yandex_session(table_id, cookies)

    access_ok, final_url = check_yandex_auth(cookies)
    if not access_ok:
        update_table_yandex_status(table_id, "auth_required", f"Нужен вход администратора в Яндекс: {final_url}")
        return {
            "status": "need_login",
            "detail": "Требуется вход администратора",
            "yandex_status": "auth_required",
            "cookies_count": len(cookies),
            "final_url": final_url,
        }

    update_table_yandex_status(table_id, "connected", None)
    return {
        "status": "ok",
        "detail": "Яндекс-сессия обновлена",
        "yandex_status": "connected",
        "cookies_count": len(cookies),
    }


@app.route("/api/yandex/connect", methods=["POST"])
def yandex_connect():
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401

    tables = query_db("SELECT id FROM table_workspaces WHERE user_id=? ORDER BY id DESC", (user["id"],))
    if not tables:
        return jsonify({"status": "need_login", "detail": "Нет доступных таблиц", "vnc_url": YANDEX_VNC_URL})

    result = try_refresh_yandex_session(tables[0]["id"], user["id"], reason="connect_button")
    if result.get("status") != "ok":
        return jsonify({"status": "need_login", "detail": result.get("detail") or "Требуется вход администратора", "vnc_url": YANDEX_VNC_URL})

    query_db(
        "UPDATE table_workspaces SET yandex_status='connected', yandex_last_error=NULL, yandex_last_checked_at=?, updated_at=? WHERE user_id=?",
        (now_iso(), now_iso(), user["id"]),
    )
    return jsonify({"status": "ok"})


@app.route("/api/tables/<int:table_id>/yandex/vnc/start", methods=["POST"])
def yandex_vnc_start(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_belongs_to_user(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    return jsonify({"vnc_url": YANDEX_VNC_URL})




@app.route("/api/tables/<int:table_id>/yandex/refresh", methods=["POST"])
def yandex_refresh(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_belongs_to_user(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404

    result = try_refresh_yandex_session(table_id, user["id"], reason="manual")
    body = {
        "status": result.get("status"),
        "detail": result.get("detail"),
        "yandex_status": result.get("yandex_status"),
        "cookies_count": result.get("cookies_count", 0),
        "vnc_url": YANDEX_VNC_URL,
    }
    return jsonify(body)


@app.route("/api/tables/<int:table_id>", methods=["DELETE"])
def delete_table(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db("SELECT id FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user["id"]), one=True)
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    query_db("DELETE FROM table_program_items WHERE table_id=?", (table_id,))
    query_db("DELETE FROM table_entries WHERE table_id=?", (table_id,))
    query_db("DELETE FROM table_workspaces WHERE id=?", (table_id,))
    shutil.rmtree(storage_for_table(user["id"], table_id), ignore_errors=True)
    return jsonify({"status": "ok"})


@app.route("/api/tables/<int:table_id>/excel", methods=["POST"])
def upload_excel(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db("SELECT id FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user["id"]), one=True)
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    excel = request.files.get("excel")
    if not excel:
        return jsonify({"detail": "Файл excel обязателен"}), 400

    base = storage_for_table(user["id"], table_id)
    dst = os.path.join(base, "meta")
    os.makedirs(dst, exist_ok=True)
    for folder in ["phonograms", "receipts", "presentations"]:
        os.makedirs(os.path.join(base, folder), exist_ok=True)

    xlsx_path = os.path.join(dst, "original.xlsx")
    excel.save(xlsx_path)

    try:
        headers, preview_rows, total_rows, sheet_name = parse_excel_preview(xlsx_path)
    except Exception as exc:
        app.logger.warning("[tables_excel_preview] table_id=%s parse_failed error=%s", table_id, str(exc))
        query_db(
            "UPDATE table_workspaces SET status='error', progress=0, excel_headers_json='[]', excel_preview_rows_json='[]', excel_total_rows=0, excel_sheet_name=NULL, updated_at=? WHERE id=?",
            (now_iso(), table_id),
        )
        return jsonify({"detail": f"Excel не распознан: {str(exc)}"}), 400

    auto_mapping, signature, auto_source = apply_mapping_templates_and_presets(user["id"], headers)

    query_db("DELETE FROM table_entries WHERE table_id=?", (table_id,))

    for idx, row in enumerate(preview_rows, start=2):
        row_data = {str(i): v for i, v in enumerate(row)}
        try:
            query_db(
                """
                INSERT INTO table_entries (table_id, row_id, number_title, fio, team, unique_key, audio_url, receipt_url, consent_url, application_file, generic_file, video_url, presentation_url, audio_local, receipt_local, consent_local, application_local, generic_local, presentation_local, row_data_json, resolved_fields_json, conflicts_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    table_id,
                    idx,
                    "",
                    "",
                    "",
                    f"preview-{idx}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    json.dumps(row_data, ensure_ascii=False),
                    json.dumps({}, ensure_ascii=False),
                    json.dumps({}, ensure_ascii=False),
                    now_iso(),
                ),
            )
        except Exception:
            app.logger.exception("[tables_excel_preview] table_id=%s row_id=%s insert_failed", table_id, idx)
            raise

    query_db(
        "UPDATE table_workspaces SET excel_headers_json=?, excel_preview_rows_json=?, excel_total_rows=?, excel_sheet_name=?, mapping_json=?, status='excel_loaded', progress=100, updated_at=? WHERE id=?",
        (
            json.dumps(headers, ensure_ascii=False),
            json.dumps(preview_rows, ensure_ascii=False),
            total_rows,
            sheet_name,
            json.dumps(auto_mapping, ensure_ascii=False),
            now_iso(),
            table_id,
        ),
    )

    app.logger.info(
        "[tables_excel_preview] table_id=%s sheet=%s headers=%s preview_rows=%s total_rows=%s",
        table_id,
        sheet_name,
        len(headers),
        len(preview_rows),
        total_rows,
    )

    can_start, reason = table_required_mapping_status(auto_mapping)

    return jsonify(
        {
            "status": "ok",
            "table_status": "excel_loaded",
            "headers": headers,
            "rows": preview_rows,
            "total_rows": total_rows,
            "sheet": sheet_name,
            "mapping": auto_mapping,
            "mapping_signature": signature,
            "mapping_autofilled": bool(auto_mapping),
            "mapping_autofill_source": auto_source,
            "can_start": can_start,
            "reason": reason,
        }
    )


@app.route("/api/tables/<int:table_id>/excel_preview", methods=["GET"])
def table_excel_preview(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db(
        "SELECT id, excel_headers_json, excel_preview_rows_json, excel_total_rows FROM table_workspaces WHERE id=? AND user_id=?",
        (table_id, user["id"]),
        one=True,
    )
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    headers = json.loads(table["excel_headers_json"] or "[]")
    rows = json.loads(table["excel_preview_rows_json"] or "[]")
    total_rows = int(table["excel_total_rows"] or 0)
    return jsonify({
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "mapping_fields": MAPPING_FIELDS,
        "mapping_signature": mapping_signature(headers),
    })


@app.route("/api/tables/<int:table_id>/excel-data", methods=["GET"])
def table_excel_data(table_id):
    return table_excel_preview(table_id)


@app.route("/api/tables/<int:table_id>/mapping", methods=["GET"])
def get_table_mapping(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db("SELECT id, mapping_json, excel_headers_json FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user["id"]), one=True)
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    mapping = normalize_mapping(json.loads(table["mapping_json"] or "{}"))
    headers = json.loads(table["excel_headers_json"] or "[]")
    can_start, reason = table_required_mapping_status(mapping)
    return jsonify({
        "mapping": mapping,
        "can_start": can_start,
        "reason": reason,
        "mapping_fields": MAPPING_FIELDS,
        "mapping_signature": mapping_signature(headers),
    })


@app.route("/api/tables/<int:table_id>/mapping", methods=["POST"])
def set_table_mapping(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db("SELECT id FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user["id"]), one=True)
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404

    payload = request.get_json(silent=True) or {}
    mapping = normalize_mapping(payload.get("mapping") if isinstance(payload, dict) else payload)
    query_db("UPDATE table_workspaces SET mapping_json=?, updated_at=? WHERE id=?", (json.dumps(mapping, ensure_ascii=False), now_iso(), table_id))
    try:
        before_stats = collect_table_entries_stats(table_id)
        stats = rebuild_entries_from_excel(table_id, user["id"], mapping, preserve_files=True, clear_existing=True)
        app.logger.info(
            "[tables_entries_rebuild] stage=mapping_saved table_id=%s total=%s with_number_title=%s with_fio=%s with_audio_local=%s with_receipt_local=%s with_presentation_local=%s",
            table_id,
            stats["total"],
            stats["with_number_title"],
            stats["with_fio"],
            stats["with_audio_local"],
            stats["with_receipt_local"],
            stats["with_presentation_local"],
        )
        log_table_entries_stats("mapping_saved", table_id, previous=before_stats)
    except Exception as exc:
        app.logger.warning("[tables_entries_rebuild] stage=mapping_saved table_id=%s skipped error=%s", table_id, str(exc))
    can_start, reason = table_required_mapping_status(mapping)
    return jsonify({"status": "ok", "mapping": mapping, "can_start": can_start, "reason": reason})


@app.route("/api/tables/<int:table_id>/mapping/remember", methods=["POST"])
def remember_table_mapping(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401

    table = query_db(
        "SELECT id, excel_headers_json, mapping_json FROM table_workspaces WHERE id=? AND user_id=?",
        (table_id, user["id"]),
        one=True,
    )
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404

    headers = json.loads(table["excel_headers_json"] or "[]")
    if not headers:
        return jsonify({"detail": "Сначала загрузите Excel"}), 400

    mapping = normalize_mapping(json.loads(table["mapping_json"] or "{}"))
    if not mapping:
        return jsonify({"detail": "Сначала назначьте хотя бы одну колонку"}), 400

    signature = mapping_signature(headers)
    save_mapping_template(user["id"], signature, mapping)

    for label in mapping.keys():
        for idx in mapped_field_indexes(mapping, label):
            if idx >= len(headers):
                continue
            header_norm = normalize_header_text(headers[idx])
            upsert_mapping_preset(user["id"], label, header_norm, "exact", priority=10)

    return jsonify({"status": "ok", "signature": signature, "saved_fields": len(mapping)})


@app.route("/api/tables/<int:table_id>/start-download", methods=["POST"])
def start_download(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    t = query_db(
        "SELECT id, status, processed_count, download_cursor_row_id, mapping_json, excel_total_rows FROM table_workspaces WHERE id=? AND user_id=?",
        (table_id, user["id"]),
        one=True,
    )
    if not t:
        return jsonify({"detail": "Таблица не найдена"}), 404

    mapping = normalize_mapping(json.loads(t["mapping_json"] or "{}"))
    can_start, reason = table_required_mapping_status(mapping)
    if not can_start:
        return jsonify({"detail": reason}), 400

    resume_statuses = {"auth_required", "paused", "downloading_partial"}
    current_status = str(t["status"] or "new").lower()
    is_resume = current_status in resume_statuses and int(t["processed_count"] or 0) > 0

    if not has_global_yandex_session():
        update_table_yandex_status(table_id, "auth_required", "Нужен вход администратора в Яндекс")
        query_db(
            "UPDATE table_workspaces SET status='auth_required', last_error=?, updated_at=? WHERE id=?",
            ("Нужен вход администратора в Яндекс", now_iso(), table_id),
        )
        return jsonify({"status": "need_login", "detail": "Нужен вход администратора в Яндекс", "vnc_url": YANDEX_VNC_URL}), 400

    update_table_yandex_status(table_id, "connected", None)
    total_rows = int(t["excel_total_rows"] or 0)
    if is_resume:
        query_db(
            "UPDATE table_workspaces SET status='downloading_partial', total_count=?, last_error=NULL, updated_at=? WHERE id=?",
            (total_rows, now_iso(), table_id),
        )
    else:
        query_db(
            "UPDATE table_workspaces SET status='downloading', total_count=?, processed_count=0, download_cursor_row_id=2, progress=0, last_error=NULL, updated_at=? WHERE id=?",
            (total_rows, now_iso(), table_id),
        )

    threading.Thread(target=process_table_download, args=(table_id, user["id"]), daemon=True).start()
    return jsonify({"status": "resumed" if is_resume else "started", "mode": "resume" if is_resume else "start"})


@app.route("/api/tables/<int:table_id>/entries", methods=["GET"])
def list_table_entries(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = query_db("SELECT id FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user["id"]), one=True)
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    rows = query_db("SELECT * FROM table_entries WHERE table_id=? ORDER BY id", (table_id,))
    return jsonify([dict(r) for r in rows])


def table_owned_or_404(table_id, user_id):
    return query_db("SELECT * FROM table_workspaces WHERE id=? AND user_id=?", (table_id, user_id), one=True)


def break_label(minutes):
    mins = int(minutes or 0)
    h = mins // 60
    m = mins % 60
    if h and m:
        return f"Перерыв — {h} ч {m} мин"
    if h:
        return f"Перерыв — {h} ч"
    return f"Перерыв — {m} минут"


def list_program_items_raw(table_id):
    return query_db("SELECT * FROM table_program_items WHERE table_id=? ORDER BY sort_index, id", (table_id,))


def resolve_table_local_path(table_id, user_id, local_value):
    raw_value = str(local_value or "").strip()
    if not raw_value:
        return "", ""

    owner_user_id = get_table_owner_id(table_id) or user_id
    normalized = raw_value.replace("\\", "/")
    table_storage = Path(storage_for_table(owner_user_id, table_id)).resolve()
    candidate = Path(normalized)
    if candidate.is_absolute():
        full_path = candidate.resolve(strict=False)
    else:
        full_path = (table_storage / candidate).resolve(strict=False)
    return raw_value, str(full_path)



def has_local_entry_file(table_id, user_id, entry_id, field_name, local_value, *, placeholder_for_audio=False):
    local_raw, full_path = resolve_table_local_path(table_id, user_id, local_value)
    if not local_raw:
        return False
    if placeholder_for_audio and local_raw.lower() == "audio.txt":
        return False
    if os.path.exists(full_path):
        return True
    app.logger.warning(
        "[tables_program_missing_local_file] table_id=%s entry_id=%s field=%s path=%s",
        table_id,
        entry_id,
        field_name,
        local_raw,
    )
    return False


def apply_program_order(table_id, ordered_ids):
    ts = now_iso()
    for idx, item_id in enumerate(ordered_ids, start=1):
        query_db(
            "UPDATE table_program_items SET sort_index=?, updated_at=? WHERE id=? AND table_id=?",
            (idx * 1000, ts, item_id, table_id),
        )
    query_db("UPDATE table_workspaces SET program_updated_at=?, updated_at=? WHERE id=?", (ts, ts, table_id))


def get_program_items_payload(table_id, user_id, mapping=None):
    if mapping is None:
        table = query_db("SELECT mapping_json FROM table_workspaces WHERE id=?", (table_id,), one=True)
        mapping = normalize_mapping(json.loads((dict(table).get("mapping_json") if table else "") or "{}"))

    visible_columns = build_visible_columns(table_id, mapping)
    rows = query_db(
        """
        SELECT
            tpi.id AS program_item_id,
            tpi.kind,
            tpi.entry_id,
            tpi.break_minutes,
            tpi.sort_index,
            te.*
        FROM table_program_items tpi
        LEFT JOIN table_entries te ON te.id = tpi.entry_id AND te.table_id = tpi.table_id
        WHERE tpi.table_id=?
        ORDER BY tpi.sort_index, tpi.id
        """,
        (table_id,),
    )
    payload = []
    display_number = 0
    for r in rows:
        item = dict(r)
        if item["kind"] == "break":
            payload.append(
                {
                    "program_item_id": item["program_item_id"],
                    "kind": "break",
                    "break_minutes": item["break_minutes"],
                    "label": break_label(item["break_minutes"]),
                }
            )
            continue

        display_number += 1
        resolved_fields = json.loads(item.get("resolved_fields_json") or "{}")
        conflicts = json.loads(item.get("conflicts_json") or "{}")

        cells = []
        has_any_conflict = False
        for field_name in visible_columns:
            cell = build_tag_cell(table_id, user_id, item, mapping, field_name, resolved_fields, conflicts)
            if cell.get("conflict"):
                has_any_conflict = True
            cells.append(cell)

        has_audio = any(c["key"] == "audio_url" and c.get("files") for c in cells)
        payload.append(
            {
                "program_item_id": item["program_item_id"],
                "kind": "entry",
                "entry_id": item.get("entry_id"),
                "display_number": display_number,
                "row_id": item.get("row_id"),
                "number_title": item.get("number_title") or "",
                "fio": item.get("fio") or "",
                "team": item.get("team") or "",
                "cells": cells,
                "resolved_fields": resolved_fields,
                "conflicts": conflicts,
                "is_problematic": (not has_audio) or has_any_conflict,
            }
        )
    return payload


def clean_name_part(text, fallback):
    raw = str(text or "").strip()
    raw = re.sub(r"\s+", " ", raw)
    raw = re.sub(r'[\/:*?"<>|]', " ", raw)
    raw = re.sub(r"\s*-\s*-+", " - ", raw)
    raw = raw.strip(" .-")
    return raw or fallback


def build_program_filename(display_number, entry, ext):
    entry_id = entry.get("id")
    title = clean_name_part(entry.get("number_title"), f"entry-{entry_id}")
    fio = clean_name_part(entry.get("fio"), f"entry-{entry_id}")
    base = f"{int(display_number):03d} - {title} - {fio}"
    base = re.sub(r"\s+", " ", base)
    base = re.sub(r"-\s*-+", "-", base)
    base = base.strip()
    return make_safe_basename(base, fallback=f"{int(display_number):03d}-entry-{entry_id}") + f".{ext}"


def get_program_entry_item(table_id, item_id, user_id):
    table = table_owned_or_404(table_id, user_id)
    if not table:
        return None, None, (jsonify({"detail": "Таблица не найдена"}), 404)
    rows = list_program_items_raw(table_id)
    display_number = 0
    target = None
    for row in rows:
        row = dict(row)
        if row["kind"] == "entry":
            display_number += 1
        if row["id"] == item_id:
            target = row
            break
    if not target:
        return None, None, (jsonify({"detail": "Элемент программы не найден"}), 404)
    if target["kind"] != "entry":
        return None, None, (jsonify({"detail": "Элемент не является выступлением"}), 400)
    entry = query_db("SELECT * FROM table_entries WHERE id=? AND table_id=?", (target["entry_id"], table_id), one=True)
    if not entry:
        return None, None, (jsonify({"detail": "Строка выступления не найдена"}), 404)
    return dict(target), {**dict(entry), "display_number": display_number}, None


@app.route("/api/tables/<int:table_id>/finalize", methods=["POST"])
def finalize_table(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = table_owned_or_404(table_id, user["id"])
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    if table is not None and not isinstance(table, dict):
        table = dict(table)

    mapping = normalize_mapping(json.loads(table.get("mapping_json") or "{}"))

    entries_stats = query_db(
        """
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN TRIM(COALESCE(number_title, '')) <> '' THEN 1 ELSE 0 END) AS with_number_title,
            SUM(CASE WHEN TRIM(COALESCE(fio, '')) <> '' THEN 1 ELSE 0 END) AS with_fio
        FROM table_entries
        WHERE table_id=?
        """,
        (table_id,),
        one=True,
    )
    total_entries = int(entries_stats["total"] or 0)
    nonempty_titles = int(entries_stats["with_number_title"] or 0)
    nonempty_fio = int(entries_stats["with_fio"] or 0)

    if total_entries == 0 or (total_entries > 0 and nonempty_titles == 0 and nonempty_fio == 0):
        try:
            before_stats = collect_table_entries_stats(table_id)
            rebuild_stats = rebuild_entries_from_excel(table_id, user["id"], mapping, preserve_files=True, clear_existing=True)
            app.logger.info(
                "[tables_entries_rebuild] stage=finalize table_id=%s total=%s with_number_title=%s with_fio=%s with_audio_local=%s with_receipt_local=%s with_presentation_local=%s",
                table_id,
                rebuild_stats["total"],
                rebuild_stats["with_number_title"],
                rebuild_stats["with_fio"],
                rebuild_stats["with_audio_local"],
                rebuild_stats["with_receipt_local"],
                rebuild_stats["with_presentation_local"],
            )
            log_table_entries_stats("finalize", table_id, previous=before_stats)
        except Exception as exc:
            return jsonify({"detail": f"Данные таблицы не подготовлены для финализации: {str(exc)}"}), 400

        entries_stats = query_db(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN TRIM(COALESCE(number_title, '')) <> '' THEN 1 ELSE 0 END) AS with_number_title,
                SUM(CASE WHEN TRIM(COALESCE(fio, '')) <> '' THEN 1 ELSE 0 END) AS with_fio
            FROM table_entries
            WHERE table_id=?
            """,
            (table_id,),
            one=True,
        )
        total_entries = int(entries_stats["total"] or 0)
        nonempty_titles = int(entries_stats["with_number_title"] or 0)
        nonempty_fio = int(entries_stats["with_fio"] or 0)

    if total_entries == 0:
        return jsonify({"detail": "Нет строк для формирования программы"}), 400
    if nonempty_titles == 0 and nonempty_fio == 0:
        return jsonify({"detail": "Данные table_entries пустые: заполните mapping и подготовьте материалы"}), 400

    existing_items = query_db("SELECT id, kind, entry_id, sort_index FROM table_program_items WHERE table_id=? ORDER BY sort_index, id", (table_id,))
    if existing_items:
        entry_ids = [r["entry_id"] for r in existing_items if r["kind"] == "entry" and r["entry_id"] is not None]
        existing_entries = query_db("SELECT id FROM table_entries WHERE table_id=?", (table_id,))
        existing_entry_ids = {r["id"] for r in existing_entries}
        missing = [eid for eid in entry_ids if eid not in existing_entry_ids]
        if missing:
            return jsonify({"detail": "Программа содержит ссылки на удаленные строки. Пересоберите программу вручную."}), 409
        query_db(
            "UPDATE table_workspaces SET is_finalized=1, finalized_at=COALESCE(finalized_at, ?), program_updated_at=COALESCE(program_updated_at, ?), updated_at=? WHERE id=?",
            (now_iso(), now_iso(), now_iso(), table_id),
        )
        return jsonify({"status": "ok", "already_finalized": True})

    entries = query_db("SELECT id FROM table_entries WHERE table_id=? ORDER BY COALESCE(row_id, id), id", (table_id,))
    ts = now_iso()
    for idx, e in enumerate(entries, start=1):
        query_db(
            "INSERT INTO table_program_items (table_id, kind, entry_id, sort_index, break_minutes, created_at, updated_at) VALUES (?, 'entry', ?, ?, NULL, ?, ?)",
            (table_id, e["id"], idx * 1000, ts, ts),
        )
    query_db(
        "UPDATE table_workspaces SET is_finalized=1, finalized_at=?, program_updated_at=?, updated_at=? WHERE id=?",
        (ts, ts, ts, table_id),
    )
    return jsonify({"status": "ok", "created_items": len(entries)})


@app.route("/api/tables/<int:table_id>/program", methods=["GET"])
def get_program(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    table = table_owned_or_404(table_id, user["id"])
    if not table:
        return jsonify({"detail": "Таблица не найдена"}), 404
    table_dict = dict(table)
    mapping = normalize_mapping(json.loads(table_dict.get("mapping_json") or "{}"))
    visible_columns = build_visible_columns(table_id, mapping)
    visible_tags = [
        {
            "key": key,
            "label": MAPPING_FIELDS.get(key, key),
            "type": tag_type_for_key(key),
        }
        for key in visible_columns
    ]
    return jsonify({
        "table_id": table_id,
        "is_finalized": int(table["is_finalized"] or 0) == 1,
        "items": get_program_items_payload(table_id, user["id"], mapping=mapping),
        "visible_tags": visible_tags,
        "tag_type_config": TAG_TYPE_OVERRIDES,
    })


@app.route("/api/tables/<int:table_id>/entries/<int:entry_id>/resolve", methods=["PATCH"])
def resolve_entry_grouped_conflict(table_id, entry_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_belongs_to_user(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    entry = query_db("SELECT id, resolved_fields_json FROM table_entries WHERE id=? AND table_id=?", (entry_id, table_id), one=True)
    if not entry:
        return jsonify({"detail": "Строка не найдена"}), 404
    payload = request.get_json(silent=True) or {}
    field = str(payload.get("field") or "").strip()
    value = str(payload.get("value") or "").strip()
    if not field:
        return jsonify({"detail": "field обязателен"}), 400
    entry_dict = dict(entry)
    resolved = json.loads(entry_dict.get("resolved_fields_json") or "{}")
    resolved[field] = value
    query_db("UPDATE table_entries SET resolved_fields_json=? WHERE id=? AND table_id=?", (json.dumps(resolved, ensure_ascii=False), entry_id, table_id))
    return jsonify({"status": "ok", "resolved_fields": resolved})


@app.route("/api/tables/<int:table_id>/program/reorder", methods=["PATCH"])
def reorder_program(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_owned_or_404(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    payload = request.get_json(silent=True) or {}
    ids = payload.get("program_item_ids") if isinstance(payload, dict) else None
    if not isinstance(ids, list) or not ids:
        return jsonify({"detail": "program_item_ids должен быть непустым списком"}), 400
    rows = list_program_items_raw(table_id)
    existing_ids = [r["id"] for r in rows]
    if sorted(existing_ids) != sorted(ids):
        return jsonify({"detail": "Список элементов не совпадает с программой"}), 400
    apply_program_order(table_id, ids)
    return jsonify({"status": "ok", "items": get_program_items_payload(table_id, user["id"])})


@app.route("/api/tables/<int:table_id>/program/item/<int:item_id>/move_to_position", methods=["PATCH"])
def move_program_item(table_id, item_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_owned_or_404(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    payload = request.get_json(silent=True) or {}
    position = int(payload.get("position") or 0)
    if position <= 0:
        return jsonify({"detail": "position должен быть больше 0"}), 400

    rows = [dict(r) for r in list_program_items_raw(table_id)]
    by_id = {r["id"]: r for r in rows}
    if item_id not in by_id:
        return jsonify({"detail": "Элемент программы не найден"}), 404
    if by_id[item_id]["kind"] != "entry":
        return jsonify({"detail": "Можно перемещать только выступления"}), 400

    full_ids = [r["id"] for r in rows]
    remaining_full = [i for i in full_ids if i != item_id]
    remaining_entries = [r["id"] for r in rows if r["kind"] == "entry" and r["id"] != item_id]
    target_pos = min(max(position, 1), len(remaining_entries) + 1)

    if target_pos > len(remaining_entries):
        insert_idx = len(remaining_full)
    else:
        target_entry_id = remaining_entries[target_pos - 1]
        insert_idx = remaining_full.index(target_entry_id)

    remaining_full.insert(insert_idx, item_id)
    apply_program_order(table_id, remaining_full)
    return jsonify({"status": "ok", "items": get_program_items_payload(table_id, user["id"])})


@app.route("/api/tables/<int:table_id>/program/break", methods=["POST"])
def add_program_break(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_owned_or_404(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    payload = request.get_json(silent=True) or {}
    minutes = int(payload.get("break_minutes") or 0)
    after_item_id = payload.get("after_item_id")
    before_item_id = payload.get("before_item_id")
    if minutes <= 0:
        return jsonify({"detail": "break_minutes должен быть больше 0"}), 400

    rows = [dict(r) for r in list_program_items_raw(table_id)]
    all_ids = [r["id"] for r in rows]

    insert_index = len(all_ids)
    if before_item_id is not None:
        try:
            before_item_id = int(before_item_id)
            insert_index = all_ids.index(before_item_id)
        except Exception:
            return jsonify({"detail": "before_item_id не найден в программе"}), 400
    elif after_item_id is not None:
        try:
            after_item_id = int(after_item_id)
            insert_index = all_ids.index(after_item_id) + 1
        except Exception:
            return jsonify({"detail": "after_item_id не найден в программе"}), 400

    ts = now_iso()
    query_db(
        "INSERT INTO table_program_items (table_id, kind, entry_id, sort_index, break_minutes, created_at, updated_at) VALUES (?, 'break', NULL, ?, ?, ?, ?)",
        (table_id, (len(rows) + 1) * 1000, minutes, ts, ts),
    )
    new_item_id = query_db("SELECT last_insert_rowid() AS id", one=True)["id"]
    all_ids.insert(insert_index, new_item_id)
    apply_program_order(table_id, all_ids)
    return jsonify({"status": "ok", "items": get_program_items_payload(table_id, user["id"])})


@app.route("/api/tables/<int:table_id>/program/item/<int:item_id>", methods=["DELETE"])
def delete_program_item(table_id, item_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_owned_or_404(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    item = query_db("SELECT * FROM table_program_items WHERE id=? AND table_id=?", (item_id, table_id), one=True)
    if not item:
        return jsonify({"detail": "Элемент программы не найден"}), 404
    if item["kind"] != "break":
        return jsonify({"detail": "Удалять можно только перерывы"}), 400
    query_db("DELETE FROM table_program_items WHERE id=? AND table_id=?", (item_id, table_id))
    rows = [r["id"] for r in list_program_items_raw(table_id)]
    apply_program_order(table_id, rows)
    return jsonify({"status": "ok", "items": get_program_items_payload(table_id, user["id"])})


def send_program_file(table_id, item_id, kind, user_id):
    item, entry, err = get_program_entry_item(table_id, item_id, user_id)
    if err:
        return err

    column = f"{kind}_local"
    local_value = entry.get(column) or ""
    part = str(request.args.get("part") or "").strip()
    local_parts = parse_local_file_values(local_value)
    selected_local = ""
    if part and part in local_parts:
        selected_local = part
    elif local_parts:
        selected_local = local_parts[0]
    elif local_value:
        selected_local = local_value

    owner_user_id = get_table_owner_id(table_id) or user_id
    base_path = os.path.join(storage_for_table(owner_user_id, table_id), selected_local) if selected_local else ""
    if kind == "audio" and (not selected_local or selected_local.lower() == "audio.txt" or not os.path.exists(base_path)):
        msg = "Фонограмма не предоставлена"
        buf = io.BytesIO(msg.encode("utf-8"))
        return send_file(buf, as_attachment=True, download_name=build_program_filename(entry["display_number"], entry, "txt"), mimetype="text/plain; charset=utf-8")

    if not selected_local or not os.path.exists(base_path):
        return jsonify({"detail": "Файл не найден"}), 404

    ext = os.path.splitext(selected_local)[1].lstrip(".").lower() or "bin"
    filename = build_program_filename(entry["display_number"], entry, ext)
    if kind == "receipt" and request.args.get("download") != "1" and ext == "pdf":
        return send_file(base_path, as_attachment=False, download_name=filename, mimetype="application/pdf")
    return send_file(base_path, as_attachment=True, download_name=filename)


@app.route("/api/tables/<int:table_id>/program/download/audio/<int:item_id>", methods=["GET"])
def download_program_audio(table_id, item_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    return send_program_file(table_id, item_id, "audio", user["id"])


@app.route("/api/tables/<int:table_id>/program/download/receipt/<int:item_id>", methods=["GET"])
def download_program_receipt(table_id, item_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    return send_program_file(table_id, item_id, "receipt", user["id"])


@app.route("/api/tables/<int:table_id>/program/download/presentation/<int:item_id>", methods=["GET"])
def download_program_presentation(table_id, item_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    return send_program_file(table_id, item_id, "presentation", user["id"])


@app.route("/api/tables/<int:table_id>/program/download_all", methods=["GET"])
def download_program_all(table_id):
    user = table_user_from_request()
    if not user:
        return jsonify({"detail": "Не авторизован"}), 401
    if not table_owned_or_404(table_id, user["id"]):
        return jsonify({"detail": "Таблица не найдена"}), 404
    items = get_program_items_payload(table_id, user["id"])
    owner_user_id = get_table_owner_id(table_id) or user["id"]
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            if item["kind"] != "entry":
                continue
            entry = query_db("SELECT * FROM table_entries WHERE id=? AND table_id=?", (item["entry_id"], table_id), one=True)
            if not entry:
                continue
            entry = dict(entry)
            display = item["display_number"]
            # audio
            audio_local = entry.get("audio_local") or ""
            if not audio_local or audio_local.lower() == "audio.txt":
                txt_name = build_program_filename(display, entry, "txt")
                zf.writestr(f"Фонограммы/{txt_name}", "Фонограмма не предоставлена")
            else:
                src = os.path.join(storage_for_table(owner_user_id, table_id), audio_local)
                if os.path.exists(src):
                    ext = os.path.splitext(audio_local)[1].lstrip(".") or "bin"
                    zf.write(src, arcname=f"Фонограммы/{build_program_filename(display, entry, ext)}")
            # receipt
            receipt_local = entry.get("receipt_local") or ""
            if receipt_local:
                src = os.path.join(storage_for_table(owner_user_id, table_id), receipt_local)
                if os.path.exists(src):
                    ext = os.path.splitext(receipt_local)[1].lstrip(".") or "bin"
                    zf.write(src, arcname=f"Квитки/{build_program_filename(display, entry, ext)}")
            # presentation
            pres_local = entry.get("presentation_local") or ""
            if pres_local:
                src = os.path.join(storage_for_table(owner_user_id, table_id), pres_local)
                if os.path.exists(src):
                    ext = os.path.splitext(pres_local)[1].lstrip(".") or "bin"
                    zf.write(src, arcname=f"Презентации/{build_program_filename(display, entry, ext)}")
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name=f"program_table_{table_id}.zip", mimetype="application/zip")


def resolve_file(entry_id, ftype):
    e = query_db("SELECT te.*, tw.user_id FROM table_entries te JOIN table_workspaces tw ON tw.id=te.table_id WHERE te.id=?", (entry_id,), one=True)
    if not e:
        return None, (jsonify({"detail": "Entry не найден"}), 404)

    col = f"{ftype}_local"
    if col not in e.keys() or not e[col]:
        return None, (jsonify({"detail": "Файл не найден"}), 404)

    p = os.path.join(storage_for_table(e["user_id"], e["table_id"]), e[col])
    if not os.path.exists(p):
        return None, (jsonify({"detail": "Файл отсутствует"}), 404)
    return p, None


@app.route("/api/files/<int:entry_id>/<ftype>", methods=["GET"])
def get_file(entry_id, ftype):
    if not table_user_from_request():
        return jsonify({"detail": "Не авторизован"}), 401
    p, err = resolve_file(entry_id, ftype)
    if err:
        return err
    return send_file(p, as_attachment=True, download_name=os.path.basename(p))


@app.route("/api/preview/<int:entry_id>/<ftype>", methods=["GET"])
def preview_file(entry_id, ftype):
    if not table_user_from_request():
        return jsonify({"detail": "Не авторизован"}), 401
    p, err = resolve_file(entry_id, ftype)
    if err:
        return err
    ext = os.path.splitext(p)[1].lower()
    if ext == ".mp3":
        return jsonify({"detail": "Для mp3 доступно только скачивание"}), 400
    if ext in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".pdf", ".txt"]:
        return send_file(p, as_attachment=False)
    return jsonify({"detail": "Предпросмотр не поддерживается"}), 400


# =====================================================
#                VK STREAMING API
# =====================================================

VK_SETTINGS_FILE = os.path.join(BASE_DIR, "vk_settings.json")
PREVIEW_PID_FILE = os.path.join(BASE_DIR, "start_vk_preview.pid")
START_VK_SCRIPT = os.path.join(BASE_DIR, "start_vk.py")  # путь к скрипту, который мы писали
VK_LOCK_TEMPLATE = "/tmp/start_vk_{stream}.lock"
DEFAULT_STREAM_NAME = os.environ.get("RTMP_STREAM_NAME", "stream")


def start_preview_process():
    """Запустить start_vk.py как отдельный процесс (режим preview)."""
    # если pid-файл есть — считаем, что уже запущено
    if os.path.exists(PREVIEW_PID_FILE):
        try:
            with open(PREVIEW_PID_FILE) as f:
                pid = int(f.read().strip())
            os.kill(pid, 0)
            return
        except Exception:
            # старый pid мёртв — продолжим и перезапишем
            pass

    # Запускаем скрипт с произвольным аргументом (preview)
    cmd = ["/usr/bin/python3", START_VK_SCRIPT, "preview"]
    try:
        p = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, preexec_fn=os.setsid)
        with open(PREVIEW_PID_FILE, "w") as f:
            f.write(str(p.pid))
    except Exception as e:
        app.logger.error("Не удалось запустить preview process: %s", e)

def stop_preview_process():
    """Остановить ранее запущенный preview процесс (если есть)."""
    if not os.path.exists(PREVIEW_PID_FILE):
        return
    try:
        with open(PREVIEW_PID_FILE) as f:
            pid = int(f.read().strip())
        # убиваем группу процессов
        try:
            os.killpg(os.getpgid(pid), signal.SIGTERM)
        except Exception:
            os.kill(pid, signal.SIGTERM)
    except Exception as e:
        app.logger.error("Ошибка при остановке preview: %s", e)
    try:
        os.remove(PREVIEW_PID_FILE)
    except:
        pass


def load_vk_settings():
    # если файла нет — возвращаем дефолт и создаём его
    default = {
        "enabled": False,
        "vk_rtmp_url": "",
        "scheduled_start": None,
        "title": "",
        "target_ids": [],
        "preview_path": os.path.join(BASE_DIR, "static", "vk_preview.jpg"),
        "show_preview": False
    }
    if not os.path.exists(VK_SETTINGS_FILE):
        try:
            os.makedirs(os.path.dirname(VK_SETTINGS_FILE), exist_ok=True)
            with open(VK_SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(default, f, ensure_ascii=False, indent=4)
        except Exception:
            pass
        return default

    try:
        with open(VK_SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            data.setdefault("show_preview", False)
            data.setdefault("preview_path", os.path.join(BASE_DIR, "static", "vk_preview.jpg"))
            return data
    except Exception:
        # если файл кривой — перезаписываем дефолтом
        try:
            with open(VK_SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(default, f, ensure_ascii=False, indent=4)
        except Exception:
            pass
        return default

def save_vk_settings(data):
    try:
        os.makedirs(os.path.dirname(VK_SETTINGS_FILE), exist_ok=True)
        with open(VK_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        # В лог можно записать, но мы возвращаем ошибку клиенту
        raise

def load_stream_targets():
    default_targets = []
    if not os.path.exists(STREAM_TARGETS_FILE):
        try:
            with open(STREAM_TARGETS_FILE, "w", encoding="utf-8") as f:
                json.dump(default_targets, f, ensure_ascii=False, indent=4)
        except Exception:
            pass
        return default_targets
    try:
        with open(STREAM_TARGETS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_targets

def save_stream_targets(targets):
    os.makedirs(os.path.dirname(STREAM_TARGETS_FILE), exist_ok=True)
    with open(STREAM_TARGETS_FILE, "w", encoding="utf-8") as f:
        json.dump(targets, f, ensure_ascii=False, indent=4)

def preview_url_for(settings):
    preview_path = settings.get("preview_path")
    if not preview_path:
        return ""
    if not os.path.isabs(preview_path):
        preview_path = os.path.join(BASE_DIR, preview_path)
    static_dir = os.path.join(BASE_DIR, "static")
    try:
        if os.path.commonpath([static_dir, preview_path]) == static_dir:
            return "/static/" + os.path.basename(preview_path)
    except ValueError:
        return ""
    return preview_path

@app.route("/vk/status", methods=["GET"])
@login_required
def vk_status():
    settings = load_vk_settings()
    settings["preview_url"] = preview_url_for(settings)
    settings["stream_url"] = resolve_stream_url()
    settings["targets"] = load_stream_targets()
    settings.setdefault("target_ids", [])
    settings.setdefault("show_preview", False)
    return jsonify(settings)

@app.route("/vk/register_key", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_register_key():
    data = request.get_json(silent=True) or {}
    vk_rtmp_url = (data.get("vk_rtmp_url") or "").strip()
    if not vk_rtmp_url:
        return jsonify({"status": "error", "message": "Ключ обязателен"}), 400
    settings = load_vk_settings()
    settings["vk_rtmp_url"] = vk_rtmp_url
    try:
        save_vk_settings(settings)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить ключ: {e}"}), 500
    return jsonify({"status": "ok"})

@app.route("/vk/public_status", methods=["GET"])
def vk_public_status():
    settings = load_vk_settings()
    return jsonify({
        "enabled": settings.get("enabled", False),
        "preview_url": preview_url_for(settings),
        "stream_url": resolve_stream_url(),
        "show_preview": settings.get("show_preview", False)
    })

@app.route("/vk/preview", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_preview_upload():
    file = request.files.get("image")
    if not file:
        return jsonify({"status": "error", "message": "Файл не выбран"}), 400

    s = load_vk_settings()
    try:
        preview_dir = os.path.join(BASE_DIR, "static")
        os.makedirs(preview_dir, exist_ok=True)
        preview_path = os.path.join(preview_dir, "vk_preview.jpg")
        file.save(preview_path)
        s["preview_path"] = preview_path
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить превью: {e}"}), 500

    return jsonify({"status": "ok", "preview_url": preview_url_for(s)})

@app.route("/vk/preview_visibility", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_preview_visibility():
    data = request.get_json(silent=True) or {}
    show_preview = bool(data.get("show_preview"))
    s = load_vk_settings()
    s["show_preview"] = show_preview
    try:
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить настройки: {e}"}), 500
    return jsonify({"status": "ok", "show_preview": s.get("show_preview", False)})

def _start_vk_process(stream_name):
    cmd = ["/usr/bin/python3", START_VK_SCRIPT, stream_name]
    run_as_nobody = bool(os.environ.get("VK_START_AS_NOBODY", "1") == "1")
    if run_as_nobody:
        cmd = ["sudo", "-u", "nobody", *cmd]

    p = subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        preexec_fn=os.setsid,
    )
    return p, cmd


@app.route("/vk/start_now", methods=["POST"])
@login_required
def vk_start_now():
    s = load_vk_settings()
    title = ""
    target_ids = []

    if request.is_json:
        payload = request.get_json(silent=True) or {}
        title = payload.get("title", "")
        target_ids = payload.get("target_ids") or []
    else:
        title = request.form.get("title", "")

    selected_target = None
    targets = load_stream_targets()
    if target_ids:
        selected_target = next((t for t in targets if t.get("id") == target_ids[0]), None)

    s["enabled"] = True
    s["scheduled_start"] = None
    if title:
        s["title"] = title
    s["target_ids"] = target_ids
    s["show_preview"] = False

    try:
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить настройки: {e}"}), 500

    stream_name = DEFAULT_STREAM_NAME
    lock_file = VK_LOCK_TEMPLATE.format(stream=stream_name)
    lock_removed = False
    try:
        if os.path.exists(lock_file):
            os.remove(lock_file)
            lock_removed = True
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось снять lock {lock_file}: {e}"}), 500

    try:
        proc, cmd = _start_vk_process(stream_name)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось запустить ретрансляцию: {e}"}), 500

    return jsonify({
        "status": "ok",
        "ok": True,
        "msg": "VK stream start requested",
        "stream_name": stream_name,
        "target": selected_target,
        "target_ids": target_ids,
        "lock_removed": lock_removed,
        "pid": proc.pid,
        "command": cmd,
    })


@app.route("/vk/stop", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_stop():
    s = load_vk_settings()
    s["enabled"] = False
    s["scheduled_start"] = None
    try:
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить настройки: {e}"}), 500
    return jsonify({"status": "ok", "msg": "VK streaming stopped"})

@app.route("/vk/schedule", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_schedule():
    s = load_vk_settings()

    title = ""
    date = ""
    time_val = ""
    target_ids = []

    # поддерживаем два варианта: multipart/form-data (с файлом) и application/json
    if request.content_type and request.content_type.startswith("application/json"):
        body = request.get_json(silent=True) or {}
        title = body.get("title", "")
        iso = body.get("iso") or body.get("scheduled_start")
        target_ids = body.get("target_ids") or []
        if iso:
            scheduled = iso
        else:
            date = body.get("date", "")
            time_val = body.get("time", "")
    else:
        # multipart/form-data или обычная форма
        try:
            title = request.form.get("title", "")
            date = request.form.get("date", "")
            time_val = request.form.get("time", "")
            target_ids = request.form.getlist("target_ids")
        except Exception:
            # если парсинг формы упал — вернём ошибку
            return jsonify({"status": "error", "message": "Ошибка при чтении формы"}), 400

    if 'scheduled' in locals():
        pass  # уже есть
    else:
        if date and time_val:
            scheduled = f"{date}T{time_val}:00"
        else:
            now = datetime.datetime.utcnow() + datetime.timedelta(minutes=1)
            scheduled = now.isoformat()

    # Превью (если пришёл файл)
    if request.files:
        file = request.files.get("image")
    else:
        file = None

    if file:
        try:
            preview_dir = os.path.join(BASE_DIR, "static")
            os.makedirs(preview_dir, exist_ok=True)
            preview_path = os.path.join(preview_dir, "vk_preview.jpg")
            file.save(preview_path)
            s["preview_path"] = preview_path
        except Exception as e:
            return jsonify({"status": "error", "message": f"Не удалось сохранить превью: {e}"}), 500

    s["title"] = title
    s["scheduled_start"] = scheduled
    s["enabled"] = True
    s["target_ids"] = target_ids
    s["show_preview"] = False

    try:
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить настройки: {e}"}), 500

    return jsonify({"status": "ok", "scheduled": scheduled})

@app.route("/vk/targets", methods=["POST"])
@login_required
@roles_required("admin", "editor")
def vk_targets():
    data = request.get_json(silent=True) or {}
    target_ids = data.get("target_ids") or []
    s = load_vk_settings()
    s["target_ids"] = target_ids
    try:
        save_vk_settings(s)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Не удалось сохранить настройки: {e}"}), 500
    return jsonify({"status": "ok"})

@app.route("/stream/targets", methods=["GET", "POST"])
@login_required
def stream_targets():
    if request.method == "GET":
        return jsonify(load_stream_targets())

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    url_val = (data.get("url") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "Название обязательно"}), 400
    targets = load_stream_targets()
    target_id = secrets.token_hex(4)
    targets.append({"id": target_id, "name": name, "url": url_val, "enabled": True})
    save_stream_targets(targets)
    return jsonify({"status": "ok", "target": {"id": target_id, "name": name, "url": url_val, "enabled": True}})

@app.route("/stream/targets/<target_id>", methods=["PUT", "DELETE"])
@login_required
def stream_targets_update(target_id):
    targets = load_stream_targets()
    target = next((t for t in targets if t.get("id") == target_id), None)
    if not target:
        return jsonify({"status": "error", "message": "Цель не найдена"}), 404

    if request.method == "DELETE":
        targets = [t for t in targets if t.get("id") != target_id]
        save_stream_targets(targets)
        return jsonify({"status": "ok"})

    data = request.get_json(silent=True) or {}
    target["name"] = (data.get("name") or target["name"]).strip()
    target["url"] = (data.get("url") or target["url"]).strip()
    if "enabled" in data:
        target["enabled"] = bool(data.get("enabled"))
    save_stream_targets(targets)
    return jsonify({"status": "ok", "target": target})

# =====================================================
#                   END VK STREAMING
# =====================================================

# ------------ Запуск ------------
init_table_header_tags()
init_db()
ensure_admin_exists()
migrate_tables_to_current_admin()
start_tables_background_jobs()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8083, debug=True)
